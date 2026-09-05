import copy
import importlib
import json
import os
import re
import sys
import threading
import time
import uuid
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path

from build_identity import load_build_identity
from company_rules import DEFAULT_COMPANY
from document_types import (
    MANUAL_REVIEW_FOLDER,
    classify_cwt_document_type as _classify_cwt_document_type,
    normalize_document_type_for_archive as _normalize_document_type_for_archive,
)
from email_channel import resolve_channel
from frontend_run_context import ensure_run_context_dirs, load_run_context, make_run_staging_dir, serialize_run_context
from glm_runtime import GlmRequestError, GlmRuntime
from report_service import ReportService
from recognition_policy import (
    CloudProviderId,
    RecognitionMode,
    RecognitionPolicy,
    RecognitionPolicyError,
)
from run_coordinator import RunCoordinator, RunDependencies, RunRequest
from run_lifecycle import RunLifecycle, RunState
from run_state_store import RunStateStore
from provider_direct_invoice import DIRECT_INVOICE_FAMILIES
from url_trace_sanitizer import (
    build_url_evidence,
    sanitize_persistence_payload,
    sanitize_url_trace_record,
    stable_hash,
)
from user_settings import (
    DEFAULT_GLM_MODEL_CANDIDATES,
    DEFAULT_GLM_PROFILE_LIMITS,
    UserSettingsStore,
    ensure_directory,
    get_default_save_path as resolve_default_save_path,
    get_default_debug_trace_path,
    get_output_state_dir,
    get_packaged_diagnostics_dir,
)


DEFAULT_TRUTH_AUDIT_FINALIZE_TIMEOUT_SECONDS = 120.0
PROVIDER_URL_RETRY_REASON_CODES = frozenset(
    {
        "URL_DOWNLOAD_FAILED",
        "URL_PAGE_TIMEOUT",
        "URL_RECOVERY_DEADLINE_EXCEEDED",
        "URL_RECOVERY_WORKER_FAILED",
    }
)


class TruthAuditTimeout(RuntimeError):
    reason_code = "TRUTH_AUDIT_TIMEOUT"
    user_message = "真值审计收尾超时，请查看诊断报告后重试。"


class ProcessingLoopFailure(RuntimeError):
    reason_code = "PROCESSING_FAILED"
    user_message = "处理过程中发生异常，请重试；如持续失败请查看诊断报告。"


class ImapLoginError(RuntimeError):
    reason_code = "IMAP_LOGIN_FAILED"
    user_message = "邮箱登录失败，请检查授权码和 IMAP 设置。"


class RemoteAuthError(RuntimeError):
    reason_code = "REMOTE_AUTH_FAILED"
    user_message = "GLM API 身份验证失败，请检查 API Key。"


class QuotaExhaustedError(RuntimeError):
    reason_code = "QUOTA_EXHAUSTED"
    user_message = "GLM API 额度已耗尽，请充值或更换可用的 API Key。"


class TruthAuditEvidenceError(RuntimeError):
    def __init__(self, reason_code, user_message):
        super().__init__(reason_code)
        self.reason_code = reason_code
        self.user_message = user_message


@dataclass(frozen=True)
class TruthAuditPaths:
    run_root: Path
    monitoring_dir: Path
    artifact_dir: Path
    report_path: Path
    error_path: Path
    index_path: Path
    run_config_path: Path


@dataclass(frozen=True)
class TruthAuditJob:
    run_id: str
    paths: TruthAuditPaths
    thread: threading.Thread
    ready: threading.Event
    abandoned: threading.Event
    result: dict


@dataclass(frozen=True)
class _RunAdmissionCandidate:
    rules_text: str
    requested_save_path: str
    effective_save_path: str
    date_from: str
    date_to: str
    account_id: str
    channel_id: str
    email_domain: str
    run_context_json: str
    started_at: str

    def run_context(self):
        return json.loads(self.run_context_json)


class _RunAdmissionSecrets:
    __slots__ = ("email_address", "auth_code", "api_key")

    def __init__(self, email_address, auth_code, api_key):
        self.email_address = str(email_address or "")
        self.auth_code = str(auth_code or "")
        self.api_key = str(api_key or "")

    def __repr__(self):
        return "_RunAdmissionSecrets(<process-only redacted>)"


class _ProcessingPipelineSession:
    def __init__(
        self,
        *,
        api,
        candidates,
        pipeline,
        archive_service,
        save_path,
        output_state_dir,
        working_history,
        business_records,
        sidecar,
        trace_store,
        owned_extractor=None,
        provider_retry_delay_seconds=0.0,
        url_recovery_scheduler=None,
    ):
        self._api = api
        self.candidates = candidates
        self._pipeline = pipeline
        self._archive_service = archive_service
        self._save_path = save_path
        self._output_state_dir = output_state_dir
        self._working_history = working_history
        self._business_records = business_records
        self._sidecar = sidecar
        self._trace_store = trace_store
        self._owned_extractor = owned_extractor
        self._provider_retry_delay_seconds = max(
            0.0, float(provider_retry_delay_seconds)
        )
        if url_recovery_scheduler is None:
            from deferred_url_recovery import (
                DEFAULT_URL_RECOVERY_MAX_WORKERS,
                DeferredUrlRecoveryScheduler,
            )

            url_recovery_scheduler = DeferredUrlRecoveryScheduler(
                max_workers=DEFAULT_URL_RECOVERY_MAX_WORKERS,
                stop_requested=lambda: bool(
                    getattr(self._api, "_stop_requested", False)
                ),
            )
        self._url_recovery_scheduler = url_recovery_scheduler
        self._provider_url_candidates = []
        self._candidate_total = len(candidates)
        self._closed = False

    def extract(self):
        candidates = list(self.candidates)
        primary = [
            candidate
            for candidate in candidates
            if candidate.identity.source_kind != "url"
        ]
        self._provider_url_candidates = [
            candidate
            for candidate in candidates
            if candidate.identity.source_kind == "url"
        ]
        return self._pipeline.extract(primary, progress_total=self._candidate_total)

    def _record_deferred_outcomes(self, outcomes):
        for outcome in outcomes:
            event = {
                "document_id": outcome.candidate.identity.document_id,
                "sequence": outcome.candidate.sequence,
                "status": outcome.status,
                "reason_code": outcome.reason_code,
            }
            self._trace_store.set_fields(
                event["document_id"], extraction_result=dict(event)
            )
            self._api._safe_emit_stage_event(
                "extraction_pipeline", "trace", dict(event)
            )

    def _recover_deferred_candidate(self, candidate):
        resolver = getattr(self._pipeline, "resolve_one", None)
        if callable(resolver):
            return resolver(candidate)
        outcomes = self._pipeline.extract([candidate], _emit_progress_events=False)
        if len(outcomes) != 1:
            raise RuntimeError("deferred recovery did not return exactly one outcome")
        return outcomes[0]

    def _retry_deferred_candidate(self, candidate):
        outcomes = self._pipeline.retry_current_run_failures([candidate])
        if len(outcomes) != 1:
            raise RuntimeError("deferred retry did not return exactly one outcome")
        return outcomes[0]

    def _canonical_info_by_document_id(self, archived_outcomes):
        document_ids = {
            archived.outcome.candidate.identity.document_id
            for archived in archived_outcomes
        }
        records = {}
        getter = getattr(self._trace_store, "get_record", None)
        if callable(getter):
            records = {
                document_id: getter(document_id) or {}
                for document_id in document_ids
            }
        else:
            iterator = getattr(self._trace_store, "iter_records", None)
            if callable(iterator):
                records = {
                    str(record.get("document_id") or ""): record
                    for record in iterator()
                    if isinstance(record, dict)
                }
        return {
            document_id: dict(record.get("normalized_fields") or {})
            for document_id, record in records.items()
            if isinstance(record, dict)
        }

    @staticmethod
    def _is_strong_provider_retry_candidate(candidate):
        legacy = candidate.to_legacy()
        return bool(
            candidate.identity.source_kind == "url"
            and legacy.get("provider_family")
        )

    def _wait_for_provider_retry(self) -> bool:
        remaining = self._provider_retry_delay_seconds
        while remaining > 0:
            if bool(getattr(self._api, "_stop_requested", False)):
                return False
            interval = min(0.25, remaining)
            time.sleep(interval)
            remaining -= interval
        return not bool(getattr(self._api, "_stop_requested", False))

    def archive(self, outcomes):
        from archive_service import ArchiveReport
        from candidate_pipeline import partition_redundant_provider_candidates

        def merge_reports(reports):
            return ArchiveReport(
                outcomes=tuple(
                    sorted(
                        (
                            archived
                            for source_report in reports
                            for archived in source_report.outcomes
                        ),
                        key=lambda archived: archived.outcome.candidate.sequence,
                    )
                ),
                archived_count=sum(item.archived_count for item in reports),
                retained_count=sum(item.retained_count for item in reports),
                manual_count=sum(item.manual_count for item in reports),
                unresolved_count=sum(item.unresolved_count for item in reports),
                duplicate_count=sum(item.duplicate_count for item in reports),
            )

        primary_outcomes = list(outcomes)
        primary_report = self._archive_service.archive(
            primary_outcomes, self._save_path, finalize=False
        )
        skipped, pending = partition_redundant_provider_candidates(
            self._provider_url_candidates,
            primary_report.outcomes,
            canonical_info_by_document_id=self._canonical_info_by_document_id(
                primary_report.outcomes
            ),
        )
        skipped_by_document_id = {
            outcome.candidate.identity.document_id: outcome for outcome in skipped
        }
        deferred_candidates = sorted(
            [
                *(outcome.candidate for outcome in skipped),
                *pending,
            ],
            key=lambda candidate: candidate.sequence,
        )

        def recover_one(candidate):
            skipped_outcome = skipped_by_document_id.get(candidate.identity.document_id)
            if skipped_outcome is not None:
                return skipped_outcome
            return self._recover_deferred_candidate(candidate)

        provider_outcomes = self._url_recovery_scheduler.recover(
            deferred_candidates,
            recover_one,
            progress_offset=len(primary_outcomes),
            progress_total=self._candidate_total,
        )
        self._record_deferred_outcomes(provider_outcomes)
        failed_provider_outcomes = [
            outcome
            for outcome in provider_outcomes
            if outcome.status == "unresolved"
        ]
        failed_document_ids = {
            outcome.candidate.identity.document_id
            for outcome in failed_provider_outcomes
        }
        archiveable_outcomes = sorted(
            [
                *(
                    outcome
                    for outcome in provider_outcomes
                    if outcome.candidate.identity.document_id
                    not in failed_document_ids
                ),
            ],
            key=lambda outcome: outcome.candidate.sequence,
        )
        reports = [primary_report]
        if archiveable_outcomes:
            reports.append(
                self._archive_service.archive(
                    archiveable_outcomes, self._save_path, finalize=False
                )
            )

        archived_evidence = tuple(
            archived
            for source_report in reports
            for archived in source_report.outcomes
        )
        recovered_failures, still_pending = partition_redundant_provider_candidates(
            [outcome.candidate for outcome in failed_provider_outcomes],
            archived_evidence,
            canonical_info_by_document_id=self._canonical_info_by_document_id(
                archived_evidence
            )
        )
        self._record_deferred_outcomes(recovered_failures)
        still_pending_ids = {
            candidate.identity.document_id for candidate in still_pending
        }
        remaining_failures = [
            outcome
            for outcome in failed_provider_outcomes
            if outcome.candidate.identity.document_id in still_pending_ids
        ]
        retryable_failures = [
            outcome
            for outcome in remaining_failures
            if outcome.reason_code in PROVIDER_URL_RETRY_REASON_CODES
            and self._is_strong_provider_retry_candidate(outcome.candidate)
        ]
        retryable_ids = {
            outcome.candidate.identity.document_id for outcome in retryable_failures
        }
        nonretryable_failures = [
            outcome
            for outcome in remaining_failures
            if outcome.candidate.identity.document_id not in retryable_ids
        ]
        retried_failures = []
        recovered_retry_failures = []
        if retryable_failures and self._wait_for_provider_retry():
            retry_outcomes = self._url_recovery_scheduler.recover(
                [outcome.candidate for outcome in retryable_failures],
                self._retry_deferred_candidate,
                emit_progress=False,
            )
            retried_failures = [
                outcome
                for outcome in retry_outcomes
                if outcome.status == "unresolved"
            ]
            retried_failure_ids = {
                outcome.candidate.identity.document_id
                for outcome in retried_failures
            }
            retried_archiveable = [
                outcome
                for outcome in retry_outcomes
                if outcome.candidate.identity.document_id not in retried_failure_ids
            ]
            if retried_archiveable:
                reports.append(
                    self._archive_service.archive(
                        retried_archiveable, self._save_path, finalize=False
                    )
                )
            archived_evidence = tuple(
                archived
                for source_report in reports
                for archived in source_report.outcomes
            )
            recovered_retry_failures, retry_still_pending = (
                partition_redundant_provider_candidates(
                    [outcome.candidate for outcome in retried_failures],
                    archived_evidence,
                    canonical_info_by_document_id=self._canonical_info_by_document_id(
                        archived_evidence
                    ),
                )
            )
            self._record_deferred_outcomes(recovered_retry_failures)
            retry_pending_ids = {
                candidate.identity.document_id for candidate in retry_still_pending
            }
            retried_failures = [
                outcome
                for outcome in retried_failures
                if outcome.candidate.identity.document_id in retry_pending_ids
            ]
        elif retryable_failures:
            retried_failures = list(retryable_failures)
        terminal_failures = sorted(
            [
                *recovered_failures,
                *nonretryable_failures,
                *recovered_retry_failures,
                *retried_failures,
            ],
            key=lambda outcome: outcome.candidate.sequence,
        )
        if terminal_failures:
            reports.append(
                self._archive_service.archive(
                    terminal_failures, self._save_path, finalize=False
                )
            )

        report = merge_reports(reports)
        report = self._archive_service.finalize(report, self._save_path)
        if not report.can_complete:
            self._api._mark_output_run_state(
                self._output_state_dir,
                "failed",
                failure_reason="processing_pipeline_incomplete",
            )
            raise ProcessingLoopFailure("PROCESSING_PIPELINE_INCOMPLETE")
        self._api._commit_output_state(
            self._output_state_dir,
            self._working_history,
            self._business_records,
        )
        return report

    def close(self):
        if self._closed:
            return
        self._closed = True
        self._sidecar.clear()
        if getattr(self._api, "_pipeline_sidecar", None) is self._sidecar:
            delattr(self._api, "_pipeline_sidecar")
        try:
            self._trace_store.flush()
        except Exception:
            pass
        close_extractor = getattr(self._owned_extractor, "close", None)
        if callable(close_extractor):
            close_extractor()


def build_processing_history_key(info, file_name, pdf_path):
    from candidate_pipeline import build_compatibility_history_key

    return build_compatibility_history_key(info, file_name, pdf_path)


def build_document_trace_id(info, file_name, pdf_path):
    import hashlib

    info = info or {}
    source_locator = str(info.get("source_url") or info.get("filepath") or pdf_path or "").strip()
    if source_locator and not info.get("is_url", False):
        source_locator = os.path.normcase(os.path.abspath(source_locator))
    evidence = {
        "processing_history_key": build_processing_history_key(info, file_name, pdf_path),
        "source_message_uid": str(info.get("email_id") or info.get("source_email_id") or "").strip(),
        "source_filename": str(file_name or "").strip(),
        "source_locator": source_locator,
        "source_kind": str(info.get("source_kind") or "").strip(),
        "source_url": str(info.get("source_url") or "").strip(),
        "provider_group_key": str(info.get("provider_group_key") or "").strip(),
    }
    canonical = json.dumps(evidence, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(canonical.encode("utf-8")).hexdigest()


def canonical_artifact_role(info_json, file_name=""):
    info_json = info_json or {}
    doc_type = str(info_json.get("Type") or "").strip().lower()
    evidence = f"{doc_type} {file_name}".lower()

    if any(token in evidence for token in ("住宿水单", "水单", "folio", "酒店账单")):
        return "hotel_folio"
    ride_evidence = any(token in evidence for token in ("打车", "滴滴", "高德", "约车", "ride"))
    itinerary_evidence = bool(info_json.get("_is_itinerary")) or any(
        token in evidence for token in ("行程单", "行程报销单", "itinerary")
    )
    if ride_evidence and itinerary_evidence:
        return "ride_itinerary"
    if ride_evidence:
        return "ride_invoice"
    if "住宿" in evidence or "hotel" in evidence:
        return "hotel_invoice"
    return "invoice"


def build_business_record_key(code, number, info_json, file_name=""):
    identity = f"{str(code or '').strip()}_{str(number or '').strip()}"
    return f"{identity}::{canonical_artifact_role(info_json, file_name)}"


def is_business_duplicate(code, number, records, info_json, file_name=""):
    if not str(code or "").strip() and not str(number or "").strip():
        return False
    records = records or {}
    role = canonical_artifact_role(info_json, file_name)
    if build_business_record_key(code, number, info_json, file_name) in records:
        return True

    legacy_key = f"{str(code or '').strip()}_{str(number or '').strip()}"
    legacy_record = records.get(legacy_key)
    if not isinstance(legacy_record, dict):
        return False
    stored_role = str(legacy_record.get("artifact_role") or "").strip()
    if not stored_role and legacy_record.get("file"):
        stored_role = canonical_artifact_role({}, legacy_record["file"])
    return stored_role == role if stored_role else role == "invoice"


def record_business_success(records, code, number, info_json, file_name, date="", amount=""):
    key = build_business_record_key(code, number, info_json, file_name)
    records[key] = {
        "file": file_name,
        "date": date or info_json.get("Date", ""),
        "amount": amount or info_json.get("Amount", ""),
        "artifact_role": canonical_artifact_role(info_json, file_name),
    }
    return key


class QuotaExceededError(RuntimeError):
    pass


def classify_cwt_document_type(info_json, info, file_name, local_cits_fast_path=False):
    return _classify_cwt_document_type(info_json, info, file_name, local_cits_fast_path)


def normalize_document_type_for_archive(info_json, file_name, cwt_classified=False):
    return _normalize_document_type_for_archive(info_json, file_name, cwt_classified)


class _FallbackDocumentTraceStore:
    """In-memory trace store used when the optional diagnostics module is absent."""

    def __init__(self, output_path=None):
        self.output_path = os.path.abspath(
            output_path or get_default_debug_trace_path()
        )
        os.makedirs(os.path.dirname(self.output_path), exist_ok=True)
        self._records = {}
        self._order = []
        self._archive_index = {}
        self._url_document_ids = set()

    def start_document(
        self,
        source_filename,
        source_path=None,
        document_id=None,
        persistence_is_url=False,
    ):
        doc_id = document_id or uuid.uuid4().hex
        if doc_id not in self._records:
            self._records[doc_id] = {
                "source_filename": source_filename,
                "document_id": doc_id,
                "extractor_raw_result": None,
                "normalized_fields": None,
                "classification_result": None,
                "naming_result": None,
                "combine_keys": None,
                "combine_result": None,
                "archive_target": None,
                "failure_reason": None,
            }
            if source_path:
                self._records[doc_id]["source_path"] = source_path
            self._order.append(doc_id)
        if persistence_is_url:
            self._url_document_ids.add(doc_id)
        return doc_id

    def get_record(self, document_id):
        return self._records.get(document_id)

    def iter_records(self):
        for document_id in self._order:
            yield self._records[document_id]

    def set_fields(self, document_id, **fields):
        record = self._records.get(document_id)
        if not record:
            return

        for key, value in fields.items():
            record[key] = copy.deepcopy(value)

        archive_target = fields.get("archive_target")
        if archive_target:
            self.bind_archive_target(document_id, archive_target)

    def bind_archive_target(self, document_id, archive_target):
        record = self._records.get(document_id)
        if not record or not archive_target:
            return

        old_target = record.get("archive_target")
        if old_target:
            self._archive_index.pop(self._normalize_path(old_target), None)

        normalized = self._normalize_path(archive_target)
        self._archive_index[normalized] = document_id
        record["archive_target"] = archive_target

    def move_archive_target(self, old_target, new_target):
        if not old_target or not new_target:
            return None

        normalized_old = self._normalize_path(old_target)
        document_id = self._archive_index.pop(normalized_old, None)
        if not document_id:
            return None

        normalized_new = self._normalize_path(new_target)
        self._archive_index[normalized_new] = document_id
        self._records[document_id]["archive_target"] = new_target
        return document_id

    def get_document_id_by_archive_target(self, archive_target):
        if not archive_target:
            return None
        return self._archive_index.get(self._normalize_path(archive_target))

    def record_failure_event(self, document_id, code, stage, message=None, severity="failure"):
        record = self._records.get(document_id)
        if not record:
            return

        event = {
            "code": code,
            "stage": stage,
            "severity": severity,
        }
        if message:
            event["message"] = message

        failure_reason = record.get("failure_reason")
        if not failure_reason:
            record["failure_reason"] = dict(event)
            record["failure_reason"]["history"] = [dict(event)]
            return

        history = failure_reason.setdefault("history", [])
        history.append(dict(event))

        if self._severity_rank(severity) >= self._severity_rank(failure_reason.get("severity")):
            failure_reason.update(event)

    def flush(self):
        temp_path = f"{self.output_path}.tmp"
        with open(temp_path, "w", encoding="utf-8") as handle:
            for document_id in self._order:
                record = self._records[document_id]
                if document_id in self._url_document_ids:
                    record = sanitize_url_trace_record(record)
                json.dump(
                    record,
                    handle,
                    ensure_ascii=False,
                    default=str,
                )
                handle.write("\n")
        os.replace(temp_path, self.output_path)

    @staticmethod
    def _normalize_path(path):
        return os.path.normcase(os.path.abspath(path))

    @staticmethod
    def _severity_rank(severity):
        ranks = {
            None: 0,
            "info": 1,
            "skipped": 2,
            "fallback": 3,
            "failure": 4,
        }
        return ranks.get(severity, 0)

class InvoiceAppAPI:
    def __init__(self, truth_audit_timeout_seconds=None, revision_resolver=None):
        self._run_context = load_run_context()
        if revision_resolver is None:
            from run_evidence import default_revision

            revision_resolver = default_revision
        self._revision_resolver = revision_resolver
        self._diag_lock = threading.Lock()
        self._url_retention_ledger_lock = threading.Lock()
        self._packaged_diag_enabled = bool(getattr(sys, "frozen", False))
        self._packaged_diag_poll_count = 0
        self._packaged_diag_last_progress_signature = None
        self._packaged_diag_excepthook_installed = False
        self.progress = 0
        self.status_text = "Ready to start..."
        self.logs = []
        self._is_running = False
        self._stop_requested = False
        self.run_state = "idle"
        self.last_error = ""
        self.quota_exhausted = False
        self.quota_message = ""
        self._worker_thread = None
        self._truth_audit_thread = None
        self._truth_audit_job = None
        if truth_audit_timeout_seconds is None:
            truth_audit_timeout_seconds = DEFAULT_TRUTH_AUDIT_FINALIZE_TIMEOUT_SECONDS
        self._truth_audit_timeout_seconds = max(0.001, float(truth_audit_timeout_seconds))
        self._admission_lock = threading.RLock()
        self._run_lifecycle = RunLifecycle()
        self._active_run_handle = None
        self._active_temp_dir = None
        self._terminal_frontend_run_id = ""
        self._settings_store = UserSettingsStore()
        self._local_llm_provider = None
        self._local_llm_provider_lock = threading.Lock()
        self._requested_save_path = ""
        self._effective_save_path = ""
        self._effective_date_from = ""
        self._effective_date_to = ""
        self._raw_date_range_display = ""
        self._imap_query_range_display = ""
        self._current_run_id = self._run_context.get("run_id", "")
        self.audit_counts = {"manual_check": 0, "retention": 0, "raw_invoices": 0}
        self.discovered_categories = set()
        self.processed_invoices = []
        self.error_invoices = []
        self.stats = {"emails": 0, "invoices": 0, "errors": 0}
        self._email_level_url_evidence_seen = set()
        self._last_export_path = ""
        self._active_run_config = {}
        self._timing_metrics = {}
        self._run_state_store = RunStateStore(state_sink=self._apply_run_state_snapshot)
        self._install_packaged_thread_excepthook()
        self.build_identity = load_build_identity()
        self._validate_runtime_build_identity()
        # Release-prep candidate: startup must not trigger local developer scans.

    def _refresh_run_context(self):
        refreshed_context = load_run_context()
        if refreshed_context != self._run_context:
            self._run_context = refreshed_context
        ensure_run_context_dirs(self._run_context)
        self._current_run_id = self._run_context.get("run_id", "")
        return self._run_context

    def _prepare_run_lifecycle(self):
        active = self._active_run_handle
        if active is not None:
            raise RuntimeError("a run handle is already assigned")

        run_token = uuid.uuid4().hex
        external_run_id = str(self._current_run_id or self._run_context.get("run_id", "") or "run")
        lifecycle_run_id = f"{external_run_id}-{run_token}"
        staging_dir = make_run_staging_dir(self._run_context, lifecycle_run_id)
        handle = self._run_lifecycle.begin(lifecycle_run_id, staging_dir)
        run_root = str(self._run_context.get("run_root", "") or "").strip()
        temp_root = Path(run_root).resolve() / "temp" if run_root else Path.cwd() / "temp"
        self._active_temp_dir = temp_root / lifecycle_run_id
        self._active_run_handle = handle
        self._terminal_frontend_run_id = ""
        return handle

    def _apply_run_state_snapshot(self, snapshot):
        previous_state = getattr(self, "run_state", "idle")
        self.progress = int(snapshot.get("progress", 0) or 0)
        self.status_text = str(snapshot.get("status_text", ""))
        self.logs = copy.deepcopy(list(snapshot.get("logs", [])))
        self.discovered_categories = set(snapshot.get("new_categories", []))
        self.stats = copy.deepcopy(dict(snapshot.get("stats", {})))
        self.processed_invoices = copy.deepcopy(list(snapshot.get("processed_invoices", [])))
        self.error_invoices = copy.deepcopy(list(snapshot.get("error_invoices", [])))
        self._is_running = bool(snapshot.get("is_running", False))
        self.run_state = str(snapshot.get("run_state", "idle"))
        self.last_error = str(snapshot.get("last_error", ""))
        self._stop_requested = bool(snapshot.get("stop_requested", False))
        self.quota_exhausted = bool(snapshot.get("quota_exhausted", False))
        self.quota_message = str(snapshot.get("quota_message", ""))
        if previous_state != self.run_state:
            self._safe_emit_run_state_event(previous_state, self.run_state)

    def _active_staging_path(self):
        if self._active_run_handle is not None:
            return str(self._active_run_handle.staging_dir)
        return str(self._run_context.get("staging_dir") or "staging")

    @staticmethod
    def _resolve_truth_audit_paths(context, run_id):
        context = dict(context or {})
        configured_root = str(context.get("run_root") or "").strip()
        if configured_root:
            run_root = Path(configured_root).resolve()
        else:
            configured_monitoring = str(context.get("monitoring_dir") or "").strip()
            run_root = (
                Path(configured_monitoring).resolve().parent
                if configured_monitoring
                else Path.cwd().resolve()
            )
        monitoring_dir = run_root / "monitoring"
        safe_run_id = re.sub(
            r"[^A-Za-z0-9._-]+",
            "-",
            str(run_id or "run"),
        ).strip("-.") or "run"
        artifact_dir = monitoring_dir / "quarantined" / safe_run_id
        return TruthAuditPaths(
            run_root=run_root,
            monitoring_dir=monitoring_dir,
            artifact_dir=artifact_dir,
            report_path=artifact_dir / "email_truth_audit.json",
            error_path=artifact_dir / "email_truth_audit_error.json",
            index_path=artifact_dir / "truth_audit_index.json",
            run_config_path=monitoring_dir / "run_config.json",
        )

    def _monitoring_path(self, filename):
        run_context = self._refresh_run_context()
        monitoring_dir = run_context.get("monitoring_dir", "")
        if not monitoring_dir:
            return ""
        return os.path.join(monitoring_dir, filename)

    def _snapshot_counts(self):
        return {
            "emails": int(getattr(self, "stats", {}).get("emails", 0) or 0),
            "archived": int(getattr(self, "stats", {}).get("invoices", 0) or 0),
            "manual_check": int(getattr(self, "audit_counts", {}).get("manual_check", 0) or 0),
            "retention": int(getattr(self, "audit_counts", {}).get("retention", 0) or 0),
            "raw_invoices": int(getattr(self, "audit_counts", {}).get("raw_invoices", 0) or 0),
            "errors": int(getattr(self, "stats", {}).get("errors", 0) or 0),
        }

    def _reset_timing_metrics(self):
        self._timing_metrics = {}

    def _record_timing_metric(self, name, elapsed_seconds):
        metric_name = str(name or "").strip()
        if not metric_name:
            return
        try:
            elapsed = max(0.0, float(elapsed_seconds or 0.0))
        except (TypeError, ValueError):
            return
        bucket = self._timing_metrics.setdefault(
            metric_name,
            {"count": 0, "total_seconds": 0.0, "max_seconds": 0.0},
        )
        bucket["count"] = int(bucket.get("count", 0) or 0) + 1
        bucket["total_seconds"] = float(bucket.get("total_seconds", 0.0) or 0.0) + elapsed
        bucket["max_seconds"] = max(float(bucket.get("max_seconds", 0.0) or 0.0), elapsed)

    def _build_timing_breakdown(self):
        breakdown = {}
        for metric_name in sorted(self._timing_metrics):
            bucket = self._timing_metrics.get(metric_name, {})
            count = int(bucket.get("count", 0) or 0)
            total_seconds = float(bucket.get("total_seconds", 0.0) or 0.0)
            max_seconds = float(bucket.get("max_seconds", 0.0) or 0.0)
            breakdown[metric_name] = {
                "count": count,
                "total_seconds": round(total_seconds, 3),
                "avg_seconds": round(total_seconds / count, 3) if count else 0.0,
                "max_seconds": round(max_seconds, 3),
            }
        return breakdown

    def _resolve_active_company(self):
        configured_company = str((self._active_run_config or {}).get("company", "") or "").strip()
        if configured_company:
            return configured_company
        return (
            self._run_context.get("company")
            or (self._settings_store.load() or {}).get("company")
            or DEFAULT_COMPANY
        )

    def _diag_append_jsonl(self, path, payload):
        if not path:
            return
        try:
            os.makedirs(os.path.dirname(path), exist_ok=True)
            with self._diag_lock:
                with open(path, "a", encoding="utf-8") as fh:
                    fh.write(json.dumps(payload, ensure_ascii=False, default=str))
                    fh.write("\n")
        except Exception as exc:
            print(f"[diag] append jsonl failed: {path}: {exc}")

    def _diag_write_json(self, path, payload):
        if not path:
            return
        try:
            os.makedirs(os.path.dirname(path), exist_ok=True)
            temp_path = f"{path}.tmp"
            with self._diag_lock:
                with open(temp_path, "w", encoding="utf-8") as fh:
                    json.dump(payload, fh, ensure_ascii=False, indent=2, default=str)
                os.replace(temp_path, path)
        except Exception as exc:
            print(f"[diag] write json failed: {path}: {exc}")

    def _packaged_diag_dir(self):
        if not self._packaged_diag_enabled:
            return ""
        try:
            return get_packaged_diagnostics_dir()
        except Exception:
            return ""

    def _packaged_diag_file(self):
        diag_dir = self._packaged_diag_dir()
        if not diag_dir:
            return ""
        return os.path.join(diag_dir, "packaged_5p_diag.jsonl")

    def _packaged_diag_email_domain(self, email_address):
        email_text = str(email_address or "").strip()
        if "@" not in email_text:
            return ""
        return email_text.split("@", 1)[1].lower()

    def _sensitive_summary(self, email_address="", auth_code="", api_key=""):
        return {
            "email_domain": self._packaged_diag_email_domain(email_address),
            "has_auth_code": bool(auth_code),
            "has_api_key": bool(api_key),
        }

    def _create_document_trace_store(self, output_path=None):
        try:
            trace_module = importlib.import_module("diagnostics_trace")
            trace_store_cls = getattr(trace_module, "DocumentTraceStore")
            return trace_store_cls(output_path=output_path)
        except Exception as exc:
            self._packaged_diag_write(
                "document_trace_store_fallback",
                "_run_processing_loop",
                "exception",
                summary={"include_traceback": False},
                exc=exc,
            )
            return _FallbackDocumentTraceStore(output_path=output_path)

    def _packaged_diag_summary(self, summary):
        if not isinstance(summary, dict):
            return {}

        allowed_keys = {
            "requested_save_path",
            "effective_save_path",
            "date_from",
            "date_to",
            "email_domain",
            "has_auth_code",
            "has_api_key",
            "email_count",
            "attachment_count",
            "file_name",
            "thread_is_alive",
            "target_progress",
            "attachments",
            "poll_index",
            "logs_count",
            "include_traceback",
            "thread_target",
            "thread_target_ident",
            "connect_result",
            "scan_result_count",
            "extract_result_count",
            "url_result_count",
        }
        allowed = {key: value for key, value in summary.items() if key in allowed_keys}
        return sanitize_persistence_payload(allowed)

    def _packaged_diag_write(self, stage, function_name, outcome, summary=None, exc=None):
        if not self._packaged_diag_enabled:
            return

        exc_type = ""
        exc_message_hash = ""
        if exc is not None:
            exc_type = getattr(getattr(exc, "__class__", None), "__name__", "") or type(exc).__name__
            exc_message_hash = stable_hash(str(exc))

        self._diag_append_jsonl(
            self._packaged_diag_file(),
            {
                "ts": time.strftime("%Y-%m-%d %H:%M:%S"),
                "stage": stage,
                "function": function_name,
                "outcome": outcome,
                "thread_name": threading.current_thread().name,
                "thread_ident": threading.current_thread().ident,
                "pid": os.getpid(),
                "run_state": self.run_state,
                "progress": self.progress,
                "status_hash": stable_hash(self.status_text),
                "summary": self._packaged_diag_summary(summary),
                "exc_type": exc_type,
                "exc_message_hash": exc_message_hash,
            },
        )

    def _packaged_diag_reset(self, summary=None):
        if not self._packaged_diag_enabled:
            return

        diag_file = self._packaged_diag_file()
        if not diag_file:
            return

        try:
            os.makedirs(os.path.dirname(diag_file), exist_ok=True)
            with self._diag_lock:
                with open(diag_file, "w", encoding="utf-8"):
                    pass
        except Exception as exc:
            print(f"[diag] reset packaged jsonl failed: {diag_file}: {exc}")
            return

        self._packaged_diag_poll_count = 0
        self._packaged_diag_last_progress_signature = None
        self._packaged_diag_write("session_start", "start_processing", "success", summary=summary)

    def _install_packaged_thread_excepthook(self):
        if not self._packaged_diag_enabled or self._packaged_diag_excepthook_installed:
            return

        previous_hook = getattr(threading, "excepthook", None)

        def _hook(args):
            try:
                thread_obj = getattr(args, "thread", None)
                self._packaged_diag_write(
                    "thread_unhandled_exception",
                    "threading.excepthook",
                    "exception",
                    summary={
                        "thread_target": getattr(thread_obj, "name", ""),
                        "thread_target_ident": getattr(thread_obj, "ident", None),
                    },
                    exc=getattr(args, "exc_value", None),
                )
            except Exception:
                pass

            if previous_hook:
                previous_hook(args)

        threading.excepthook = _hook
        self._packaged_diag_excepthook_installed = True

    def _packaged_diag_log_progress_poll(self, payload):
        if not self._packaged_diag_enabled:
            return

        self._packaged_diag_poll_count += 1
        signature = (
            payload.get("progress"),
            payload.get("run_state"),
            payload.get("status_text"),
            payload.get("last_error"),
        )
        should_log = self._packaged_diag_poll_count <= 20 or signature != self._packaged_diag_last_progress_signature
        self._packaged_diag_last_progress_signature = signature
        if not should_log:
            return

        self._packaged_diag_write(
            "progress_poll",
            "get_progress",
            "success",
            summary={
                "poll_index": self._packaged_diag_poll_count,
                "logs_count": len(payload.get("logs", []) or []),
            },
        )

    def _safe_emit_run_state_event(self, from_state, to_state):
        if not self._run_context.get("enabled"):
            return
        self._diag_append_jsonl(
            self._monitoring_path("run_state_events.jsonl"),
            {
                "ts": time.strftime("%Y-%m-%d %H:%M:%S"),
                "from_state": from_state,
                "to_state": to_state,
                "progress": self.progress,
                "status_text": self.status_text,
                "last_error": self.last_error,
                **self._snapshot_counts(),
            },
        )

    def _safe_emit_stage_event(self, stage, event, extra=None):
        if not self._run_context.get("enabled"):
            return
        payload = {
            "ts": time.strftime("%Y-%m-%d %H:%M:%S"),
            "stage": stage,
            "event": event,
            "run_state": self.run_state,
            "progress": self.progress,
            "status_hash": stable_hash(self.status_text),
            **self._snapshot_counts(),
        }
        if extra:
            safe_extra = sanitize_persistence_payload(dict(extra))
            handle = self._active_run_handle
            if (
                handle is not None
                and handle.state is RunState.FAILED
                and safe_extra.get("result") in {"completed", "stopped", "stopped_before_extract", "stopped_after_extract"}
            ):
                safe_extra["result"] = "failed"
                safe_extra["reason"] = handle.error
            payload.update(safe_extra)
        self._diag_append_jsonl(self._monitoring_path("stage_events.jsonl"), payload)

    def _safe_emit_artifact_event(self, kind, path, document_id=None, source_kind=None, reason_code=None, category=None, extra=None):
        if not self._run_context.get("enabled"):
            return
        if source_kind == "url":
            payload = {
                "ts": time.strftime("%Y-%m-%d %H:%M:%S"),
                "artifact_id": stable_hash(
                    {"kind": kind, "path": path, "document_id": document_id}
                ),
                "path_hash": stable_hash(path),
                **self._snapshot_counts(),
            }
            payload.update(
                sanitize_persistence_payload(
                    {
                        "kind": kind,
                        "source_kind": source_kind,
                        "reason_code": reason_code,
                        "category": category,
                    }
                )
            )
            if document_id:
                payload["document_hash"] = stable_hash(document_id)
            if extra:
                safe_extra = sanitize_persistence_payload(dict(extra))
                payload.update(
                    {
                        key: value
                        for key, value in safe_extra.items()
                        if key in {
                            "status",
                            "type",
                            "category",
                            "provider",
                            "provider_family",
                            "reason_code",
                        }
                        or key.endswith(("_count", "_ms", "_seconds"))
                    }
                )
            self._diag_append_jsonl(
                self._monitoring_path("artifact_events.jsonl"), payload
            )
            return
        payload = {
            "ts": time.strftime("%Y-%m-%d %H:%M:%S"),
            "kind": kind,
            "path": path,
            "document_id": document_id,
            "source_kind": source_kind,
            "reason_code": reason_code,
            "category": category,
            **self._snapshot_counts(),
        }
        if extra:
            payload.update(extra)
        self._diag_append_jsonl(self._monitoring_path("artifact_events.jsonl"), payload)

    def _safe_emit_input_inventory_event(self, payload):
        if not self._run_context.get("enabled"):
            return
        self._diag_append_jsonl(
            self._monitoring_path("input_attachment_inventory.jsonl"),
            {
                "ts": time.strftime("%Y-%m-%d %H:%M:%S"),
                **payload,
            },
        )

    def _attachment_diag_metadata(self, info, file_name=None, document_id=None, extra=None):
        payload = {
            "document_id": document_id,
            "email_id": info.get("email_id"),
            "sender": info.get("sender"),
            "subject": info.get("subject", ""),
            "file_name": file_name or info.get("file_name") or os.path.basename(str(info.get("filepath", ""))),
            "original_filename": info.get("original_filename"),
            "attachment_ext": info.get("attachment_ext"),
            "payload_size": info.get("payload_size"),
            "mime_content_type": info.get("mime_content_type"),
            "content_disposition": info.get("content_disposition"),
            "attachment_pair_key": info.get("attachment_pair_key"),
            "sibling_pdf_present": info.get("sibling_pdf_present"),
            "sibling_ofd_present": info.get("sibling_ofd_present"),
            "sibling_xml_present": info.get("sibling_xml_present"),
            "provider_unzipped_pair_suspected": info.get("provider_unzipped_pair_suspected"),
            "zip_context": info.get("zip_context"),
            "candidate_bucket": info.get("candidate_bucket"),
            "candidate_action": info.get("candidate_action"),
            "candidate_index": info.get("candidate_index"),
            "source_kind": info.get("source_kind"),
            "prefilter_reason_code": info.get("prefilter_reason_code"),
            "prefilter_signals": info.get("prefilter_signals", {}),
            "source_url": info.get("source_url"),
            "resolved_url": info.get("resolved_url"),
            "anchor_text": info.get("anchor_text"),
            "url_host": info.get("url_host"),
            "url_path": info.get("url_path"),
            "provider_family": info.get("provider_family"),
            "provider_expected_fields": info.get("provider_expected_fields"),
            "provider_group_key": info.get("provider_group_key"),
            "provider_candidate_urls": info.get("provider_candidate_urls"),
            "provider_recovered_fields": info.get("provider_recovered_fields"),
            "provider_recovery_status": info.get("provider_recovery_status"),
            "download_mode": info.get("download_mode"),
            "wrapper_detected": info.get("wrapper_detected"),
        }
        if extra:
            payload.update(extra)
        return self._sanitize_url_persistence_payload(payload)

    @staticmethod
    def _sanitize_url_persistence_payload(payload):
        return sanitize_persistence_payload(dict(payload or {}))

    def _is_controlled_truth_run(self):
        if not self._run_context.get("enabled"):
            return False

        run_id = str(self._run_context.get("run_id", "") or "").lower()
        if any(token in run_id for token in ("lockcheck", "regression", "frontend_full_run", "manual_frontend_run")):
            return True

        return bool(self._run_context.get("monitoring_dir"))

    def _build_email_level_url_evidence_key(self, info):
        if not info or not info.get("is_url", False):
            return ""

        provider_family = str(info.get("provider_family") or "").strip().lower()
        if provider_family:
            return ""

        email_id = str(info.get("email_id") or "").strip()
        subject = str(info.get("subject") or "").strip().lower()
        sender = str(info.get("sender") or "").strip().lower()
        host = str(info.get("url_host") or "").strip().lower()

        if email_id:
            return f"email:{email_id}"
        if subject or sender:
            return f"fallback:{sender}|{subject}"
        if host:
            return f"host:{host}"
        return ""

    def _should_capture_email_level_url_evidence(self, info):
        aggregation_key = self._build_email_level_url_evidence_key(info)
        if not aggregation_key:
            return True, ""
        if aggregation_key in self._email_level_url_evidence_seen:
            return False, aggregation_key
        self._email_level_url_evidence_seen.add(aggregation_key)
        return True, aggregation_key

    def _effective_save_dir(self, requested_save_path):
        if self._run_context.get("enabled"):
            return self._run_context.get("output_dir", requested_save_path)
        return requested_save_path

    def _effective_date_range(self, date_from, date_to):
        if not self._run_context.get("enabled"):
            return date_from, date_to
        locked_from = self._run_context.get("locked_date_from", "") or date_from
        locked_to = self._run_context.get("locked_date_to", "") or date_to
        return locked_from, locked_to

    def _safe_write_run_config(
        self,
        email_address,
        auth_code="",
        api_key="",
        *,
        request=None,
    ):
        if not self._run_context.get("enabled"):
            return
        lifecycle_run_id = str(
            self._active_run_handle.run_id if self._active_run_handle is not None else ""
        )
        canonical_run_id = str(
            lifecycle_run_id
            or getattr(request, "run_id", "")
            or self._current_run_id
        )
        paths = self._resolve_truth_audit_paths(self._run_context, canonical_run_id)
        from run_evidence import RevisionUnavailable, default_hardware

        hardware_mode, hardware_fingerprint = default_hardware()
        before_exclusive = str(getattr(request, "before_exclusive", "") or "")
        account_channel = str(getattr(request, "channel_id", "") or "")
        mailbox = str(getattr(request, "mailbox", "") or "INBOX")
        run_mode = str(getattr(request, "run_mode", "") or "controlled-run")
        target_identifier = str(
            getattr(request, "target_identifier", "")
            or (self._active_run_config or {}).get("company")
            or ""
        )
        candidate_version = str(
            getattr(request, "candidate_version", "") or "source"
        )
        candidate_revision = str(
            getattr(request, "trusted_revision", "") or ""
        ).strip().lower()
        if not re.fullmatch(r"[0-9a-f]{40}", candidate_revision):
            raise RevisionUnavailable()
        self._diag_write_json(
            str(paths.run_config_path),
            {
                **serialize_run_context(self._run_context),
                "run_id": canonical_run_id,
                "staging_dir": self._active_staging_path(),
                "monitoring_dir": str(paths.monitoring_dir),
                "truth_audit_artifact_dir": str(paths.artifact_dir),
                "truth_audit_index_path": str(paths.index_path),
                "email_domain": self._packaged_diag_email_domain(email_address),
                "has_auth_code": bool(auth_code),
                "has_api_key": bool(api_key),
                "requested_save_path": self._requested_save_path,
                "effective_save_path": self._effective_save_path,
                "date_from": self._effective_date_from,
                "date_to": self._effective_date_to,
                "active_run_config": dict(self._active_run_config or {}),
                "before_exclusive": before_exclusive,
                "account_channel": account_channel,
                "mailbox": mailbox,
                "target_company": target_identifier,
                "run_mode": run_mode,
                "hardware_mode": hardware_mode,
                "hardware_fingerprint": hardware_fingerprint,
                "candidate_revision": candidate_revision,
                "candidate_version": candidate_version,
                "controlled_run": True,
                "stage_map": {
                    "start_processing": "active",
                    "run_coordinator": "active",
                    "candidate_pipeline": "active",
                    "extraction_pipeline": "active",
                    "archive_service": "active",
                    "cleanup_finalize": "active",
                },
            },
        )

    def _start_truth_audit_async(self, email_address, auth_code):
        if not self._run_context.get("enabled") or not email_address or not auth_code:
            self._truth_audit_thread = None
            self._truth_audit_job = None
            return

        context = dict(self._run_context)
        run_id = str(
            (self._active_run_handle.run_id if self._active_run_handle is not None else "")
            or self._current_run_id
            or context.get("run_id", "")
            or uuid.uuid4().hex
        )
        paths = self._resolve_truth_audit_paths(context, run_id)
        ready = threading.Event()
        abandoned = threading.Event()
        result = {}
        date_from = self._effective_date_from
        date_to = self._effective_date_to

        def _write_evidence(source_path, payload, status, reason_code, user_message):
            self._diag_write_json(str(source_path), payload)
            index_payload = {
                "schema_version": 1,
                "run_id": run_id,
                "status": status,
                "reason_code": reason_code,
                "artifact_path": str(source_path),
            }
            self._diag_write_json(str(paths.index_path), index_payload)
            valid = False
            try:
                persisted_index = json.loads(paths.index_path.read_text(encoding="utf-8"))
                persisted_artifact = Path(persisted_index.get("artifact_path", "")).resolve()
                persisted_artifact.relative_to(paths.artifact_dir.resolve())
                valid = bool(
                    persisted_index.get("run_id") == run_id
                    and persisted_index.get("status") == status
                    and persisted_index.get("reason_code") == reason_code
                    and persisted_artifact == source_path.resolve()
                    and persisted_artifact.exists()
                )
            except Exception:
                valid = False
            result.update(
                {
                    "valid": valid,
                    "status": status,
                    "reason_code": reason_code,
                    "user_message": user_message,
                    "source_path": source_path,
                    "index_path": paths.index_path,
                }
            )

        def _runner():
            try:
                module = importlib.import_module("audit_email_truth")
                collect_truth_table = getattr(module, "collect_truth_table")

                report = collect_truth_table(
                    email_address,
                    auth_code,
                    date_from,
                    date_to,
                )
                if not isinstance(report, dict):
                    raise ValueError("TRUTH_AUDIT_EVIDENCE_INVALID")
                _write_evidence(
                    paths.report_path,
                    report,
                    "success",
                    "TRUTH_AUDIT_OK",
                    "",
                )
            except ModuleNotFoundError as exc:
                _write_evidence(
                    paths.error_path,
                    {
                        "ts": time.strftime("%Y-%m-%d %H:%M:%S"),
                        "status": "skipped",
                        "reason": "AUDIT_EMAIL_TRUTH_MODULE_MISSING",
                        "error_type": type(exc).__name__,
                        "error_hash": stable_hash(str(exc))[:12],
                    },
                    "failed",
                    "AUDIT_EMAIL_TRUTH_MODULE_MISSING",
                    "真值审计模块不可用，请查看诊断报告。",
                )
            except Exception as exc:
                reason_code = (
                    "TRUTH_AUDIT_EVIDENCE_INVALID"
                    if str(exc) == "TRUTH_AUDIT_EVIDENCE_INVALID"
                    else "TRUTH_AUDIT_FAILED"
                )
                user_message = (
                    "真值审计证据无效，请查看诊断报告。"
                    if reason_code == "TRUTH_AUDIT_EVIDENCE_INVALID"
                    else "真值审计失败，请查看诊断报告。"
                )
                _write_evidence(
                    paths.error_path,
                    {
                        "ts": time.strftime("%Y-%m-%d %H:%M:%S"),
                        "reason": reason_code,
                        "error_type": type(exc).__name__,
                        "error_hash": stable_hash(str(exc))[:12],
                    },
                    "failed",
                    reason_code,
                    user_message,
                )
            finally:
                ready.set()

        thread = threading.Thread(target=_runner, name="InvoiceFlowTruthAudit", daemon=True)
        job = TruthAuditJob(
            run_id=run_id,
            paths=paths,
            thread=thread,
            ready=ready,
            abandoned=abandoned,
            result=result,
        )
        self._truth_audit_thread = thread
        self._truth_audit_job = job
        try:
            thread.start()
        except Exception:
            abandoned.set()
            self._truth_audit_thread = None
            self._truth_audit_job = None
            raise
        return thread

    def _await_truth_audit(self):
        job = self._truth_audit_job
        if job is None:
            return
        if job.thread is threading.current_thread():
            raise RuntimeError("truth audit cannot join its own thread")
        # The accepted strict audit takes about 5-12 seconds locally; 120 seconds
        # leaves a conservative production margin while still bounding finalization.
        deadline = time.monotonic() + self._truth_audit_timeout_seconds

        def _expire_job():
            job.abandoned.set()
            self._truth_audit_thread = None
            if self._truth_audit_job is job:
                self._truth_audit_job = None
            raise TruthAuditTimeout("TRUTH_AUDIT_TIMEOUT")

        if not job.ready.wait(self._truth_audit_timeout_seconds):
            _expire_job()

        active_handle = self._active_run_handle
        if (
            job.abandoned.is_set()
            or active_handle is None
            or active_handle.run_id != job.run_id
        ):
            job.abandoned.set()
            self._truth_audit_thread = None
            if self._truth_audit_job is job:
                self._truth_audit_job = None
            raise TruthAuditEvidenceError(
                "TRUTH_AUDIT_EVIDENCE_INVALID",
                "真值审计证据无效，请查看诊断报告。",
            )

        if time.monotonic() > deadline:
            _expire_job()
        self._truth_audit_thread = None
        if self._truth_audit_job is job:
            self._truth_audit_job = None
        if not job.result.get("valid"):
            raise TruthAuditEvidenceError(
                "TRUTH_AUDIT_EVIDENCE_INVALID",
                "真值审计证据无效，请查看诊断报告。",
            )
        if job.result.get("status") != "success":
            raise TruthAuditEvidenceError(
                job.result.get("reason_code") or "TRUTH_AUDIT_FAILED",
                job.result.get("user_message") or "真值审计失败，请查看诊断报告。",
            )

    def _inspect_pdf_health(self, pdf_path):
        if not pdf_path or not str(pdf_path).lower().endswith(".pdf"):
            return None

        pdf_health = {
            "exists": False,
            "size_bytes": 0,
            "starts_with_pdf_magic": False,
            "fitz_open_ok": False,
            "page_count": 0,
            "first_page_text_len": 0,
            "render_to_base64_ok": None,
            "pdf_health_class": "ok",
            "pdf_health_reason": "",
        }
        try:
            pdf_health["exists"] = os.path.exists(pdf_path)
            if not pdf_health["exists"]:
                pdf_health["pdf_health_class"] = "unopenable_pdf"
                pdf_health["pdf_health_reason"] = "missing_file"
                return pdf_health

            pdf_health["size_bytes"] = os.path.getsize(pdf_path)
            if pdf_health["size_bytes"] <= 0:
                pdf_health["pdf_health_class"] = "empty_pdf"
                pdf_health["pdf_health_reason"] = "zero_byte_file"
                return pdf_health

            try:
                with open(pdf_path, "rb") as fh:
                    pdf_health["starts_with_pdf_magic"] = fh.read(5).startswith(b"%PDF")
            except Exception:
                pass

            try:
                import fitz

                with fitz.open(pdf_path) as doc:
                    pdf_health["fitz_open_ok"] = True
                    pdf_health["page_count"] = len(doc)
                    if len(doc) > 0:
                        text_parts = []
                        for page_index in range(min(2, len(doc))):
                            text_parts.append(doc.load_page(page_index).get_text("text") or "")
                        pdf_health["first_page_text_len"] = len("".join(text_parts).strip())
            except Exception as exc:
                pdf_health["fitz_open_ok"] = False
                pdf_health["pdf_health_class"] = "corrupt_pdf" if not pdf_health["starts_with_pdf_magic"] else "unopenable_pdf"
                pdf_health["pdf_health_reason"] = str(exc)
                return pdf_health

            if pdf_health["page_count"] == 0:
                pdf_health["pdf_health_class"] = "empty_pdf"
                pdf_health["pdf_health_reason"] = "no_pages"
            return pdf_health
        except Exception as exc:
            return {
                **pdf_health,
                "pdf_health_class": "unopenable_pdf",
                "pdf_health_reason": str(exc),
            }

    def _apply_render_health(self, pdf_health, base64_img):
        if not pdf_health:
            return None
        pdf_health = dict(pdf_health)
        pdf_health["render_to_base64_ok"] = bool(base64_img)
        if base64_img:
            return pdf_health
        if pdf_health["pdf_health_class"] == "ok":
            if pdf_health.get("page_count", 0) == 0 or pdf_health.get("size_bytes", 0) == 0:
                pdf_health["pdf_health_class"] = "empty_pdf"
                pdf_health["pdf_health_reason"] = pdf_health.get("pdf_health_reason") or "no_renderable_pages"
            elif pdf_health.get("first_page_text_len", 0) == 0:
                pdf_health["pdf_health_class"] = "empty_pdf"
                pdf_health["pdf_health_reason"] = pdf_health.get("pdf_health_reason") or "empty_text_and_render_failed"
            else:
                pdf_health["pdf_health_class"] = "render_failed_pdf"
                pdf_health["pdf_health_reason"] = pdf_health.get("pdf_health_reason") or "base64_render_failed"
        return pdf_health

    @staticmethod
    def _compact_text(value):
        import re

        return re.sub(r"\s+", "", str(value or "")).strip().lower()

    @staticmethod
    def _compact_entity_text(value):
        import re
        import unicodedata

        text = unicodedata.normalize("NFKC", str(value or ""))
        return re.sub(r"\s+", "", text).strip().lower()

    def _extract_pdf_preview_text(self, pdf_path, max_pages=2):
        if not pdf_path or not str(pdf_path).lower().endswith(".pdf") or not os.path.exists(pdf_path):
            return ""
        try:
            import fitz

            texts = []
            with fitz.open(pdf_path) as doc:
                for page_index in range(min(max_pages, len(doc))):
                    texts.append(doc.load_page(page_index).get_text("text") or "")
            return "\n".join(texts).strip()
        except Exception:
            return ""

    def _has_structured_invoice_anchor(self, text):
        compact = self._compact_text(text)
        if not compact:
            return False
        anchors = [
            "发票号码",
            "发票代码",
            "购买方名称",
            "销售方名称",
            "价税合计",
            "开票日期",
            "invoice_number",
            "invoicenumber",
            "seller",
            "purchaser",
        ]
        return any(anchor in compact for anchor in anchors)

    def _match_wrapper_or_utility_reason(self, preview_text, info):
        compact = self._compact_text(preview_text)
        provider_family = str(info.get("provider_family", "") or "").lower()
        if not compact:
            return ""

        baiwang_hits = sum(
            1
            for token in ["发票预览", "下载pdf文件", "下载ofd文件", "下载xml文件", "关于百望", "previewinvoice"]
            if self._compact_text(token) in compact
        )
        if provider_family == "baiwang" and baiwang_hits >= 2 and not self._has_structured_invoice_anchor(preview_text):
            return "BAIWANG_WRAPPER_PAGE"

        if "connectingtotheitunesstore" in compact and "ifyoudonthaveitunes" in compact:
            return "UTILITY_PAGE_ITUNES_REDIRECT"

        if "票通云" in preview_text and (
            "提升运营效率" in preview_text or "发票链接一键" in preview_text or "多维链接" in preview_text
        ):
            return "UTILITY_PAGE_PROVIDER_MARKETING"

        return ""

    def _looks_like_train_ticket(self, doc_type, seller, info_json, info, file_name, pdf_path=""):
        from document_types import looks_like_train_ticket

        return looks_like_train_ticket(
            doc_type,
            seller,
            info_json,
            info,
            file_name,
            preview_loader=lambda: self._extract_pdf_preview_text(
                pdf_path, max_pages=1
            ),
        )

    def _provider_fields_match(self, expected_fields, normalized_snapshot, info_json, recovered_fields=None):
        expected = dict(expected_fields or {})
        if not any(str(value or "").strip() for value in expected.values()):
            return True, "no_expected_fields"

        normalized_snapshot = normalized_snapshot or {}
        recovered_fields = recovered_fields or {}
        invoice_number = str(
            normalized_snapshot.get("InvoiceNumber")
            or info_json.get("InvoiceNumber")
            or info_json.get("invoice_number")
            or recovered_fields.get("invoice_number")
            or ""
        ).strip()
        seller = self._compact_entity_text(
            normalized_snapshot.get("Seller")
            or info_json.get("Seller")
            or recovered_fields.get("seller")
            or ""
        )
        amount = str(
            normalized_snapshot.get("Amount")
            or info_json.get("Amount")
            or recovered_fields.get("amount")
            or ""
        ).strip()
        date = str(
            normalized_snapshot.get("Date")
            or info_json.get("Date")
            or recovered_fields.get("invoice_date")
            or ""
        ).strip()

        expected_number = str(expected.get("invoice_number") or "").strip()
        if expected_number:
            return invoice_number == expected_number, "invoice_number"

        expected_seller = self._compact_entity_text(expected.get("seller") or "")
        expected_amount = str(expected.get("amount") or "").strip()
        expected_date = str(expected.get("invoice_date") or "").strip()

        seller_match = not expected_seller or expected_seller in seller or seller in expected_seller
        amount_match = not expected_amount or expected_amount == amount
        date_match = not expected_date or expected_date == date
        return seller_match and amount_match and date_match, "seller_amount_date"

    def _evaluate_document_acceptance(self, info, info_json, normalized_snapshot, pdf_health, pdf_path):
        preview_text = self._extract_pdf_preview_text(pdf_path)
        wrapper_reason = self._match_wrapper_or_utility_reason(preview_text, info)
        provider_family = str(info.get("provider_family", "") or "").lower()
        expected_fields = info.get("provider_expected_fields") or {}
        recovered_fields = info.get("provider_recovered_fields") or {}

        result = {
            "accepted": True,
            "reason_code": "",
            "bucket": "",
            "message": "",
            "provider_family": provider_family,
            "expected_fields": expected_fields,
            "pdf_preview_excerpt": preview_text[:500],
            "pdf_health_class": (pdf_health or {}).get("pdf_health_class", ""),
        }

        if wrapper_reason:
            result.update({
                "accepted": False,
                "reason_code": wrapper_reason,
                "bucket": "provider_wrapper_rejected",
                "message": "Downloaded result still looks like a provider wrapper or utility page.",
            })
            return result

        if provider_family == "baiwang":
            matched, matched_on = self._provider_fields_match(
                expected_fields,
                normalized_snapshot,
                info_json,
                recovered_fields=recovered_fields,
            )
            result["matched_on"] = matched_on
            if not matched:
                result.update({
                    "accepted": False,
                    "reason_code": "BAIWANG_EXPECTED_ENTITY_MISMATCH",
                    "bucket": "provider_entity_mismatch",
                    "message": "Downloaded Baiwang result does not match seller/amount/date/invoice anchors from email body.",
                })
                return result

        if provider_family in DIRECT_INVOICE_FAMILIES:
            invoice_number = str(
                normalized_snapshot.get("InvoiceNumber")
                or info_json.get("InvoiceNumber")
                or info_json.get("invoice_number")
                or recovered_fields.get("invoice_number")
                or ""
            ).strip()
            expected_number = str(expected_fields.get("invoice_number") or "").strip()
            if expected_number and invoice_number and invoice_number != expected_number:
                result.update({
                    "accepted": False,
                    "reason_code": "DIRECT_INVOICE_EXPECTED_ENTITY_MISMATCH",
                    "bucket": "provider_entity_mismatch",
                    "message": "Downloaded direct-invoice PDF exposes a conflicting invoice number.",
                    "matched_on": "invoice_number_conflict",
                })
                return result

            seller = self._compact_entity_text(
                normalized_snapshot.get("Seller")
                or info_json.get("Seller")
                or recovered_fields.get("seller")
                or ""
            )
            expected_seller = self._compact_entity_text(expected_fields.get("seller") or "")
            if expected_seller and seller and expected_seller not in seller and seller not in expected_seller:
                result.update({
                    "accepted": False,
                    "reason_code": "DIRECT_INVOICE_EXPECTED_ENTITY_MISMATCH",
                    "bucket": "provider_entity_mismatch",
                    "message": "Downloaded direct-invoice PDF exposes a conflicting seller entity.",
                    "matched_on": "seller_conflict",
                })
                return result

        return result

    def _set_run_state(self, run_state, status_text=None, progress=None, last_error=None):
        previous_state = self.run_state
        self.run_state = run_state
        self._is_running = run_state in {"running", "finalizing"}
        if status_text is not None:
            self.status_text = status_text
        if progress is not None:
            self.progress = progress
        if last_error is not None:
            self.last_error = last_error
        self._safe_emit_run_state_event(previous_state, run_state)

    def _sync_run_state_store_from_legacy(self):
        self._run_state_store.update(
            progress=self.progress,
            status_text=self.status_text,
            run_state=self.run_state,
            last_error=self.last_error,
            stop_requested=self._stop_requested,
            quota_exhausted=self.quota_exhausted,
            quota_message=self.quota_message,
            statistics=self.stats,
            processed_invoices=self.processed_invoices,
            error_invoices=self.error_invoices,
            categories=self.discovered_categories,
            logs=self.logs,
        )

    def _begin_run(self, status_text):
        handle = self._active_run_handle
        if handle is not None and handle.state in {RunState.COMPLETED, RunState.FAILED}:
            self._active_run_handle = None
            handle = None
        if handle is None:
            handle = self._prepare_run_lifecycle()
        if handle.state is RunState.CREATED:
            handle.advance(RunState.SCANNING)
        self.progress = 0
        self.logs = []
        self._stop_requested = False
        self.quota_exhausted = False
        self.quota_message = ""
        self._last_export_path = ""
        self.discovered_categories.clear()
        self.processed_invoices = []
        self.error_invoices = []
        self.stats = {"emails": 0, "invoices": 0, "errors": 0}
        self.audit_counts = {"manual_check": 0, "retention": 0, "raw_invoices": 0}
        self._email_level_url_evidence_seen = set()
        self._reset_timing_metrics()
        self.build_identity = load_build_identity()
        self._set_run_state("running", status_text=status_text, progress=0, last_error="")
        build_label = self.build_identity.get("build_label", "")
        source_revision = self.build_identity.get("source_revision", "")
        review_folder = self.build_identity.get("manual_review_folder", MANUAL_REVIEW_FOLDER)
        build_stamp = " / ".join([part for part in [build_label, source_revision] if part])
        if build_stamp:
            self._append_log("构建", f"当前构建：{build_stamp}；复核目录：{review_folder}", "text-slate-400")

    def _append_log(self, level, message, color="text-slate-700"):
        self.logs.append({
            "time": time.strftime("[%H:%M:%S]"),
            "type": level,
            "color": color,
            "msg": message,
        })

    @staticmethod
    def _url_candidate_label(info):
        info = info or {}
        identity = {
            "source_url": info.get("source_url") or info.get("filepath") or "",
            "email_id": info.get("email_id") or info.get("source_email_id") or "",
            "provider_family": info.get("provider_family") or "",
        }
        return f"URL-candidate-{stable_hash(identity)[:12]}"

    def _sanitize_url_candidate_logs(self, start_index, info):
        label = self._url_candidate_label(info)
        for entry in self.logs[max(0, int(start_index or 0)):]:
            if isinstance(entry, dict):
                entry["msg"] = f"URL candidate processing event [{label}]"

    def _on_email_fetcher_progress(self, message):
        message = str(message or "").strip()
        if not message:
            return
        self.logs.append({
            "time": time.strftime("[%H:%M:%S]"),
            "type": "运行",
            "color": "text-blue-400",
            "msg": message,
        })

    @staticmethod
    def _user_safe_source_reference(source_path):
        if not source_path:
            return ""
        text = str(source_path)
        if text.startswith(("http://", "https://")):
            return text
        text = text.rstrip("\\/")
        return os.path.basename(text) or text

    def _request_safe_stop(self, message="正在安全停止，当前文件处理完后结束..."):
        if self._stop_requested:
            return
        self._stop_requested = True
        self.status_text = message
        self._append_log("停止", message, "text-amber-600")
        self._sync_run_state_store_from_legacy()

    def _resolve_quota_message(self, error_text):
        normalized = str(error_text or "").lower()
        if not normalized:
            return ""
        quota_patterns = [
            "status code 402",
            "402 client error",
            "payment required",
            "余额不足",
            "额度不足",
            "quota",
            "insufficient balance",
            "insufficient_quota",
            "billing",
        ]
        if any(token in normalized for token in quota_patterns):
            return "GLM API 额度已耗尽，请充值或更换可用的 API Key。"
        return ""

    def _mark_quota_exhausted(self, message):
        self.quota_exhausted = True
        self.quota_message = message or "GLM API 额度已耗尽，请充值或更换可用的 API Key。"
        self.status_text = self.quota_message
        self._append_log("额度", self.quota_message, "text-rose-600")

    @staticmethod
    def _safe_failure_contract(status_text, error_message):
        status = str(status_text or "")
        error = str(error_message or "")
        normalized = f"{status} {error}".upper()
        if "MISSING_REQUIRED_CREDENTIALS" in normalized:
            return "MISSING_REQUIRED_CREDENTIALS", "缺少必要凭证，请填写邮箱、授权码和 API Key。"
        if "IMAP_LOGIN_FAILED" in normalized or "邮箱登录失败" in status:
            return "IMAP_LOGIN_FAILED", "邮箱登录失败，请检查授权码和 IMAP 设置。"
        if "QUOTA_EXHAUSTED" in normalized or "额度" in status:
            return "QUOTA_EXHAUSTED", "GLM API 额度已耗尽，请充值或更换可用的 API Key。"
        if "UNRESOLVED_MAILBOX_INPUT" in normalized:
            return "UNRESOLVED_MAILBOX_INPUT", "部分邮件在重试后仍无法读取，本次任务已失败；已读取产物保留供诊断。"
        if "MAILBOX_SCAN_FAILED" in normalized:
            return "MAILBOX_SCAN_FAILED", "邮箱扫描响应异常，本次任务已失败；请重试并查看诊断报告。"
        if "WORKER_START_FAILED" in normalized or "启动失败" in status:
            return "WORKER_START_FAILED", "后台任务启动失败，请重试。"
        return "PROCESSING_FAILED", "处理过程中发生异常，请重试；如持续失败请查看诊断报告。"

    def _finish_run(self, success, status_text, last_error=""):
        handle = self._active_run_handle or self._prepare_run_lifecycle()
        if handle.state not in {RunState.COMPLETED, RunState.FAILED}:
            if not success:
                reason_code, user_message = self._safe_failure_contract(status_text, last_error)
                handle.fail(
                    RuntimeError(last_error or status_text or reason_code),
                    reason_code=reason_code,
                    user_message=user_message,
                )
            self._mark_finalizing()
            self._start_async_finalizers()

        if self._terminal_frontend_run_id == handle.run_id:
            return
        self._terminal_frontend_run_id = handle.run_id
        effective_success = bool(success and handle.state is RunState.COMPLETED)
        if effective_success:
            self._set_run_state("completed", status_text=status_text, progress=100, last_error="")
        else:
            if success:
                self.logs[:] = [entry for entry in self.logs if entry.get("type") != "完成"]
            failed_progress = self.progress if self.progress and self.progress < 100 else 99
            lifecycle_error = handle.error or last_error
            failed_status = status_text if not success else "处理失败"
            self._set_run_state("failed", status_text=failed_status, progress=failed_progress, last_error=lifecycle_error)
        self._worker_thread = None


    def _auto_start_local_scan(self):
        self.logs.append({
            "time": time.strftime("[%H:%M:%S]"),
            "type": "INFO",
            "color": "text-slate-500",
            "msg": "Local auto scan is disabled in the release-prep candidate.",
        })
        return {"success": False, "message": "AUTO_LOCAL_SCAN_DISABLED"}

    def get_default_save_path(self):
        """前端初始化时获取默认的保存路径（桌面下的发票整理文件夹）"""
        run_context = self._refresh_run_context()
        if run_context.get("enabled"):
            return ensure_directory(run_context.get("output_dir", ""))
        return ensure_directory(resolve_default_save_path())

    def _normalize_user_save_path(self, path_value=""):
        run_context = self._refresh_run_context()
        if run_context.get("enabled"):
            return ensure_directory(run_context.get("output_dir", ""))

        candidate_path = str(path_value or "").strip()
        if not candidate_path:
            candidate_path = resolve_default_save_path()
        return ensure_directory(candidate_path)

    def _output_state_dir(self, save_path):
        return get_output_state_dir(save_path or self.get_default_save_path())

    def _history_file_path(self, output_state_dir):
        return os.path.join(output_state_dir, ".antigravity_history.json")

    def _run_state_file_path(self, output_state_dir):
        return os.path.join(output_state_dir, "run_state.json")

    def _read_json_file(self, path, default):
        if not path or not os.path.exists(path):
            return copy.deepcopy(default)
        try:
            with open(path, "r", encoding="utf-8") as handle:
                return json.load(handle)
        except Exception:
            return copy.deepcopy(default)

    def _write_json_file(self, path, payload):
        os.makedirs(os.path.dirname(path), exist_ok=True)
        temp_path = f"{path}.tmp"
        with open(temp_path, "w", encoding="utf-8") as handle:
            json.dump(payload, handle, ensure_ascii=False, indent=2, default=str)
        os.replace(temp_path, path)

    def _mark_output_run_state(self, output_state_dir, status, **extra):
        payload = {
            "status": status,
            "run_id": self._current_run_id or "",
            "save_path": self._effective_save_path or self._requested_save_path or "",
            "date_from": self._effective_date_from or "",
            "date_to": self._effective_date_to or "",
            "updated_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        }
        payload.update({key: value for key, value in extra.items() if value not in (None, "")})
        self._write_json_file(self._run_state_file_path(output_state_dir), payload)
        return payload

    def _load_output_run_state(self, output_state_dir):
        return self._read_json_file(self._run_state_file_path(output_state_dir), {})

    def _load_committed_history(self, output_state_dir):
        run_state = self._load_output_run_state(output_state_dir)
        if str(run_state.get("status") or "").strip().lower() != "completed":
            return set()
        history = self._read_json_file(self._history_file_path(output_state_dir), [])
        if not isinstance(history, list):
            return set()
        return {
            str(item).strip()
            for item in history
            if isinstance(item, str) and str(item).strip()
        }

    def _commit_output_state(self, output_state_dir, history_keys, business_records):
        committed_history = sorted({str(item).strip() for item in (history_keys or set()) if str(item).strip()})
        self._write_json_file(self._history_file_path(output_state_dir), committed_history)
        self._write_json_file(os.path.join(output_state_dir, "processed_records.json"), business_records or {})
        self._mark_output_run_state(
            output_state_dir,
            "completed",
            history_count=len(committed_history),
            business_record_count=len(business_records or {}),
        )

    def _validate_runtime_build_identity(self):
        identity = dict(self.build_identity or {})
        issues = []
        if getattr(sys, "frozen", False) and not identity.get("identity_path"):
            issues.append("missing build identity payload")
        if identity.get("manual_review_folder") and identity.get("manual_review_folder") != MANUAL_REVIEW_FOLDER:
            issues.append(
                f"manual review folder mismatch: expected {MANUAL_REVIEW_FOLDER}, got {identity.get('manual_review_folder')}"
            )
        if issues:
            raise RuntimeError("Build identity validation failed: " + "; ".join(issues))

    def get_env_config(self):
        stored = self._settings_store.load() or {}
        return {
            "success": True,
            "email": str(stored.get("email", "") or ""),
            "auth_code": str(stored.get("auth_code", "") or ""),
            "api_key": str(stored.get("api_key", "") or ""),
        }

    def _default_user_settings(self):
        defaults = {
            "email": "",
            "auth_code": "",
            "api_key": "",
            "save_path": self.get_default_save_path(),
            "date_from": "",
            "date_to": "",
            "quick_range": "last_30_days",
            "company": "",
            "glm_profile_limits": copy.deepcopy(DEFAULT_GLM_PROFILE_LIMITS),
            "glm_model_candidates": copy.deepcopy(DEFAULT_GLM_MODEL_CANDIDATES),
            "remember_settings": True,
        }
        defaults.update(RecognitionPolicy().to_settings())
        return defaults

    def _get_or_create_local_llm_provider(self, policy):
        from local_llm_provider import LocalLLMProvider

        with self._local_llm_provider_lock:
            provider = self._local_llm_provider
            if (
                provider is not None
                and provider.configured_model_source == policy.local_model_source
            ):
                provider.max_tokens = policy.local_model_max_tokens
                return provider
            provider = LocalLLMProvider(
                policy.local_model_source,
                max_tokens=policy.local_model_max_tokens,
                event_sink=lambda event, payload: self._safe_emit_stage_event(
                    "local_llm", event, dict(payload)
                ),
            )
            self._local_llm_provider = provider
            return provider

    def load_user_settings(self):
        self._refresh_run_context()
        defaults = self._default_user_settings()
        stored = self._settings_store.load()

        merged = dict(defaults)
        merged.update({key: value for key, value in (stored or {}).items() if key in merged})
        merged.update(RecognitionPolicy.from_settings(merged).to_settings())
        merged["save_path"] = self._normalize_user_save_path(merged.get("save_path", ""))

        return {
            "success": True,
            "settings": merged,
            "settings_path": self._settings_store.settings_path,
        }

    def save_user_settings(self, settings):
        incoming = dict(settings or {})
        existing = self._settings_store.load() or {}
        merged = self._default_user_settings()
        remember_settings = incoming.get("remember_settings", True)
        if remember_settings:
            merged.update({key: value for key, value in existing.items() if key in merged})
            merged.update({key: value for key, value in incoming.items() if key in merged})
            merged["remember_settings"] = True
        else:
            merged["remember_settings"] = False
        merged.update(RecognitionPolicy.from_settings(merged).to_settings())
        merged["save_path"] = self._normalize_user_save_path(merged.get("save_path", ""))
        self._settings_store.save(merged)
        return {"success": True, "message": "设置已保存", "path": self._settings_store.settings_path}

    def clear_user_settings(self):
        self._settings_store.clear()
        return {"success": True, "message": "本地设置已清除"}

    def get_run_context(self):
        return serialize_run_context(self._refresh_run_context())

    def test_connection(self, email, auth_code, api_key):

        """前端测试连接时调用（真实发包到大模型验证 Key）"""
        print(
            "Testing connection",
            {
                "email_domain": self._packaged_diag_email_domain(email),
                "has_auth_code": bool(auth_code),
                "api_key_length": len(api_key),
            },
        )
        
        # 邮箱连接测试（仅在填写授权码时执行）
        if auth_code:
            try:
                from email_fetcher import EmailFetcher
                channel = resolve_channel(email)
                fetcher = EmailFetcher(email, auth_code, imap_server=channel["imap_host"])
                if not fetcher.connect():
                    return {"success": False, "message": "邮箱 IMAP 登录验证失败"}
                fetcher.disconnect()
            except Exception as e:
                return {"success": False, "message": f"邮箱连接异常: {str(e)[:50]}"}
        
        if len(api_key) <= 5:
            return {"success": False, "message": "连接失败 - API Key 格式不正确"}

        runtime = None
        try:
            payload = {
                "messages": [{"role": "user", "content": "Hi"}],
                "max_tokens": 5
            }
            runtime = GlmRuntime(api_key, settings=self._settings_store.load() or {})
            runtime.request(
                "text",
                payload,
                lambda body: body["choices"][0]["message"]["content"],
                attempts=1,
                timeout_seconds=15,
            )
            return {"success": True, "message": "连接成功 - 智谱 GLM 服务已就绪"}
        except GlmRequestError as exc:
            if exc.http_status == 401:
                return {"success": False, "message": "连接失败 - API Key 鉴权未通过或无效"}
            if exc.http_status == 402:
                return {"success": False, "message": "连接失败 - GLM API 额度已耗尽，请充值或更换 API Key"}
            if exc.http_status == 429 or exc.business_code == 1302 or exc.reason == "rate_limited":
                return {"success": False, "message": "连接失败 - 触发限流，请求并发过高"}
            if exc.reason == "timeout":
                return {"success": False, "message": "API连接失败 - 请求超时，请检查您的网络连接"}
            if exc.reason == "connection_error":
                return {"success": False, "message": "API连接失败 - 无法连接到智谱 API 服务器"}
            if exc.http_status is not None:
                return {"success": False, "message": f"连接异常 - 状态码: {exc.http_status}"}
            return {"success": False, "message": "网络或API未知异常: GLM 请求失败"}
        except Exception:
            return {"success": False, "message": "网络或API未知异常: GLM 请求失败"}
        finally:
            if runtime is not None:
                close_runtime = getattr(runtime, "close", None)
                if callable(close_runtime):
                    close_runtime()

    def test_email_auth(self, email_address, auth_code):
        if not email_address or not auth_code:
            return {"success": False, "message": "请先填写邮箱地址和授权码"}
        import imaplib
        try:
            channel = resolve_channel(email_address)
            mail = imaplib.IMAP4_SSL(channel["imap_host"])
            mail.login(email_address, auth_code)
            mail.logout()
            return {"success": True, "message": "邮箱授权验证成功"}
        except Exception:
            return {"success": False, "message": "邮箱授权验证失败，请检查授权码、IMAP 设置或网络"}



    def _processing_worker(self, request, run_handle, dependencies):
        try:
            if run_handle is None or run_handle is not self._active_run_handle:
                raise RuntimeError("reserved run handle ownership is required")
            if request.run_id != run_handle.run_id:
                raise RuntimeError("run request does not match reserved handle")
            coordinator = RunCoordinator(
                self._run_lifecycle,
                self._run_state_store,
                dependencies,
            )
            return coordinator.run(request, handle=run_handle)
        except Exception as exc:
            self._close_run_dependencies(dependencies)
            self._packaged_diag_write(
                "coordinator_worker_exception",
                "_processing_worker",
                "exception",
                exc=exc,
            )
            if run_handle is not None and run_handle.state in {RunState.COMPLETED, RunState.FAILED}:
                return None
            if run_handle is not None:
                self._finalize_admission_failure(
                    run_handle,
                    exc,
                    reason_code="COORDINATOR_START_FAILED",
                )
            return None
        finally:
            current = threading.current_thread()
            with self._admission_lock:
                if self._worker_thread is current:
                    self._worker_thread = None

    def _build_run_dependencies(
        self,
        request,
        *,
        email_address,
        auth_code,
        api_key,
        recognition_policy=None,
    ):
        from datetime import datetime, timedelta
        import inspect

        from email_fetcher import EmailFetcher
        from invoice_extractor import InvoiceExtractor
        from run_evidence import RunEvidenceWriter

        recognition_policy = recognition_policy or RecognitionPolicy.from_settings({})
        resources = {
            "fetcher": None,
            "pipeline": None,
            "recognition_policy": recognition_policy,
        }

        account_label = f"{request.channel_id}:{request.account_id}"

        def sanitize_runtime_message(message):
            text = str(message or "")
            replacements = {
                str(email_address or ""): account_label,
                str(auth_code or ""): "[redacted-auth]",
                str(api_key or ""): "[redacted-api-key]",
            }
            for sensitive, replacement in replacements.items():
                if sensitive:
                    text = re.sub(re.escape(sensitive), replacement, text, flags=re.IGNORECASE)
            return text

        def connect(_request):
            channel = resolve_channel(email_address)
            fetcher = EmailFetcher(
                email_address,
                auth_code,
                imap_server=channel["imap_host"],
                staging_dir=self._active_staging_path(),
                monitoring_dir=self._run_context.get("monitoring_dir"),
                progress_callback=lambda message: self._on_email_fetcher_progress(
                    sanitize_runtime_message(message)
                ),
            )
            resources["fetcher"] = fetcher
            self._append_log(
                "运行",
                f"正在连接邮箱通道 {request.channel_id}（账户 {request.account_id}）...",
                "text-blue-400",
            )
            if not fetcher.connect():
                raise ImapLoginError("IMAP_LOGIN_FAILED")
            return fetcher

        def scan(fetcher, _request):
            since_date = request.date_from or (
                datetime.now() - timedelta(days=30)
            ).strftime("%Y-%m-%d")
            before_date = (
                datetime.strptime(request.date_to, "%Y-%m-%d") + timedelta(days=1)
            ).strftime("%Y-%m-%d") if request.date_to else (
                datetime.now() + timedelta(days=1)
            ).strftime("%Y-%m-%d")
            self._raw_date_range_display = f"{since_date} -> {request.date_to or since_date}"
            self._imap_query_range_display = f"SINCE {since_date} / BEFORE {before_date}"
            self._coordinator_since_date = since_date
            self._coordinator_before_date = before_date
            self._append_log("运行", f"原始用户范围：{self._raw_date_range_display}", "text-blue-400")
            self._append_log("运行", f"实际 IMAP 查询：{self._imap_query_range_display}", "text-blue-400")
            email_ids = fetcher.fetch_emails_by_date(
                since_date=since_date,
                before_date=before_date,
            )
            self.stats["emails"] = len(email_ids)
            self._append_log("信息", f"共匹配到 {len(email_ids)} 封邮件。", "text-emerald-400")
            if self._stop_requested:
                self._sync_run_state_store_from_legacy()
            return email_ids

        def candidate(fetcher_email_ids, _request):
            fetcher = resources["fetcher"]
            self._append_log("运行", "正在下载并提取附件...", "text-blue-400")
            attachments = fetcher.extract_attachments(fetcher_email_ids)
            self._append_log("信息", f"共提取到 {len(attachments)} 个附件。", "text-emerald-400")
            return attachments

        def extract(attachments, _request):
            compatibility_hook = self.__dict__.get("_run_processing_loop")
            if callable(compatibility_hook):
                resources["compatibility_report"] = compatibility_hook(
                    attachments,
                    api_key,
                    request.save_path,
                    getattr(self, "_coordinator_since_date", request.date_from),
                    getattr(self, "_coordinator_before_date", request.date_to),
                    request.rules_text,
                )
                self._sync_run_state_store_from_legacy()
                return [resources["compatibility_report"]]
            extractor_parameters = inspect.signature(InvoiceExtractor).parameters.values()
            supports_glm_settings = any(
                parameter.name == "glm_settings"
                or parameter.kind == inspect.Parameter.VAR_KEYWORD
                for parameter in extractor_parameters
            )
            extractor_kwargs = {"api_key": api_key, "output_dir": request.save_path}
            if supports_glm_settings:
                extractor_kwargs["glm_settings"] = self._settings_store.load() or {}
            extractor = InvoiceExtractor(**extractor_kwargs)
            pipeline = self._create_processing_pipeline_session(
                attachments,
                api_key,
                request.save_path,
                getattr(self, "_coordinator_since_date", request.date_from),
                getattr(self, "_coordinator_before_date", request.date_to),
                request.rules_text,
                _extractor=extractor,
                _owned_extractor=extractor,
                _recognition_policy=recognition_policy,
            )
            resources["pipeline"] = pipeline
            return pipeline.extract()

        def archive(outcomes, _request):
            if "compatibility_report" in resources:
                if self.quota_exhausted:
                    raise QuotaExhaustedError(self.quota_message or "QUOTA_EXHAUSTED")
                report = resources["compatibility_report"]
                if report is None:
                    report = type(
                        "CompatibilityArchiveReport",
                        (),
                        {"can_complete": True, "archived_count": 0},
                    )()
                return report
            pipeline = resources["pipeline"]
            try:
                report = pipeline.archive(outcomes)
            except ProcessingLoopFailure:
                statuses = {str(getattr(outcome, "status", "")) for outcome in outcomes}
                if "quota_exhausted" in statuses or self.quota_exhausted:
                    raise QuotaExhaustedError("QUOTA_EXHAUSTED") from None
                if "auth_failed" in statuses:
                    raise RemoteAuthError("REMOTE_AUTH_FAILED") from None
                raise
            if self.quota_exhausted:
                raise QuotaExhaustedError(self.quota_message or "QUOTA_EXHAUSTED")
            self._sync_run_state_store_from_legacy()
            return report

        def report_callback(_context, _result):
            self._await_truth_audit()

        def pipeline_close():
            try:
                pipeline = resources.get("pipeline")
                if pipeline is not None:
                    pipeline.close()
            finally:
                self._sync_run_state_store_from_legacy()

        def disconnect_callback(_context, fetcher):
            fetcher.disconnect()

        def cleanup_callback(context):
            self._cleanup_temp_folders(
                staging_dir=context.staging_dir,
                temp_dir=self._active_temp_dir,
            )

        return RunDependencies(
            connect=connect,
            scan=scan,
            candidate=candidate,
            extract=extract,
            archive=archive,
            report_service=ReportService(
                report_callback=report_callback,
                disconnect_callback=disconnect_callback,
                cleanup_callback=cleanup_callback,
                timeout_seconds=self._truth_audit_timeout_seconds + 1.0,
                evidence_writer=RunEvidenceWriter(
                    version_resolver=lambda: request.candidate_version,
                ),
                evidence_required=request.evidence_required,
            ),
            cancel_requested=lambda: bool(self._stop_requested),
            secrets={"auth_code": auth_code, "api_key": api_key},
            finalizer_session=lambda: resources.get("fetcher"),
            pipeline_close=pipeline_close,
            state_flush=self._sync_run_state_store_from_legacy,
        )


    def _run_processing_loop(
        self,
        attachments_info,
        api_key,
        save_path,
        since_date=None,
        before_date=None,
        rules_text="",
    ):
        import inspect
        from invoice_extractor import InvoiceExtractor

        if not attachments_info:
            return self._run_processing_loop_with_extractor(
                attachments_info,
                api_key,
                save_path,
                since_date,
                before_date,
                rules_text,
                _extractor=None,
            )
        extractor_parameters = inspect.signature(InvoiceExtractor).parameters.values()
        supports_glm_settings = any(
            parameter.name == "glm_settings"
            or parameter.kind == inspect.Parameter.VAR_KEYWORD
            for parameter in extractor_parameters
        )
        extractor_kwargs = {"api_key": api_key, "output_dir": save_path}
        if supports_glm_settings:
            extractor_kwargs["glm_settings"] = self._settings_store.load() or {}
        owned_extractor = InvoiceExtractor(**extractor_kwargs)
        try:
            return self._run_processing_loop_with_extractor(
                attachments_info,
                api_key,
                save_path,
                since_date,
                before_date,
                rules_text,
                _extractor=owned_extractor,
            )
        finally:
            close_extractor = getattr(owned_extractor, "close", None)
            if callable(close_extractor):
                close_extractor()

    def _run_processing_loop_with_extractor(
        self,
        attachments_info,
        api_key,
        save_path,
        since_date=None,
        before_date=None,
        rules_text="",
        _extractor=None,
        _worker_extractor_factory=None,
        _archive_operation=None,
        _pairing_finalizer=None,
    ):
        from app_archive_adapter import AppArchiveAdapter
        from archive_service import ArchiveService
        from candidate_pipeline import CandidatePipeline, CandidatePreflight
        from extraction_pipeline import ExtractionPipeline, SharedRuntimeRemoteExtractor

        del AppArchiveAdapter, ArchiveService, CandidatePipeline, CandidatePreflight
        del ExtractionPipeline, SharedRuntimeRemoteExtractor
        session = self._create_processing_pipeline_session(
            attachments_info,
            api_key,
            save_path,
            since_date,
            before_date,
            rules_text,
            _extractor=_extractor,
            _worker_extractor_factory=_worker_extractor_factory,
            _archive_operation=_archive_operation,
            _pairing_finalizer=_pairing_finalizer,
        )
        try:
            outcomes = session.extract()
            return session.archive(outcomes)
        except (ProcessingLoopFailure, QuotaExceededError, RemoteAuthError):
            raise
        except Exception as exc:
            raise ProcessingLoopFailure("PROCESSING_PIPELINE_EXCEPTION") from exc
        finally:
            session.close()

    def _create_processing_pipeline_session(
        self,
        attachments_info,
        api_key,
        save_path,
        since_date=None,
        before_date=None,
        rules_text="",
        *,
        _extractor=None,
        _worker_extractor_factory=None,
        _archive_operation=None,
        _pairing_finalizer=None,
        _owned_extractor=None,
        _recognition_policy=None,
        _local_provider=None,
    ):
        from app_archive_adapter import AppArchiveAdapter
        from archive_service import ArchiveService
        from bounded_url_recovery import BoundedUrlRecoveryClient
        from candidate_pipeline import CandidatePipeline, CandidatePreflight
        from deferred_url_recovery import (
            DEFAULT_URL_RECOVERY_MAX_WORKERS,
            DeferredUrlRecoveryScheduler,
        )
        from extraction_pipeline import ExtractionPipeline, SharedRuntimeRemoteExtractor
        from invoice_extractor import InvoiceExtractor

        candidates = CandidatePipeline().collect(attachments_info)
        output_state_dir = self._output_state_dir(save_path)
        business_records = _extractor.load_processed_records() if _extractor else {}
        working_history = set(self._load_committed_history(output_state_dir))
        sidecar = {}
        sidecar_lock = threading.Lock()
        self._pipeline_sidecar = sidecar
        trace_store = self._create_document_trace_store(
            output_path=self._run_context.get("debug_trace_path") or None
        )
        for candidate in candidates:
            trace_store.start_document(
                source_filename=candidate.source_filename,
                source_path=candidate.source_path,
                document_id=candidate.identity.document_id,
                persistence_is_url=candidate.identity.source_kind == "url",
            )
        self._mark_output_run_state(output_state_dir, "running")

        worker_factory = _worker_extractor_factory
        if worker_factory is None:
            def worker_factory(runtime):
                return InvoiceExtractor(
                    api_key=api_key,
                    output_dir=save_path,
                    glm_runtime=runtime,
                    close_glm_runtime=False,
                )

        text_extractor = None
        result_validator = None
        prepare_remote_images = True
        if (
            _recognition_policy is not None
            and _recognition_policy.uses_local_recognition
        ):
            from local_text_extractor import LocalTextExtractor

            text_extractor = LocalTextExtractor()
            prepare_remote_images = (
                _recognition_policy.mode is not RecognitionMode.LOCAL
            )
        if _recognition_policy is not None:
            from recognition_validation import RecognitionResultValidator

            result_validator = RecognitionResultValidator(
                confidence_threshold=_recognition_policy.local_confidence_threshold
            )

        preflight = CandidatePreflight(
            api=self,
            extractor=_extractor,
            working_history=working_history,
            sidecar=sidecar,
            sidecar_lock=sidecar_lock,
            converter_factory=lambda: BoundedUrlRecoveryClient(
                staging_dir=self._active_staging_path(), timeout_ms=30000
            ),
            text_extractor=text_extractor,
            resolved_validator=result_validator,
            continue_after_validation_failure=_recognition_policy is not None,
            prepare_remote_images=prepare_remote_images,
        )
        remote = SharedRuntimeRemoteExtractor(
            owner_extractor=_extractor,
            sidecar=sidecar,
            sidecar_lock=sidecar_lock,
            worker_factory=worker_factory,
            custom_rules=rules_text,
            since_date=since_date,
            before_date=before_date,
        )
        recognizer = remote
        if _recognition_policy is not None:
            from recognition_router import (
                LocalEvidenceRecognitionExtractor,
                ModeAwareRecognitionExtractor,
            )

            def _prepared_artifact_path(candidate):
                with sidecar_lock:
                    prepared = sidecar.get(candidate.identity.document_id, {})
                    return str(prepared.get("pdf_path") or candidate.source_path)

            local_recognizer = None
            if _recognition_policy.uses_local_recognition:
                if _local_provider is None:
                    _local_provider = self._get_or_create_local_llm_provider(
                        _recognition_policy
                    )
                local_recognizer = LocalEvidenceRecognitionExtractor(
                    provider=_local_provider,
                    owner_extractor=_extractor,
                    sidecar=sidecar,
                    sidecar_lock=sidecar_lock,
                    artifact_path_resolver=_prepared_artifact_path,
                    result_validator=result_validator,
                )

            def _validate_cloud_result(outcome):
                if result_validator is None:
                    return outcome
                provider = (
                    _recognition_policy.cloud_provider.value
                    if _recognition_policy.cloud_provider is not None
                    else ""
                )
                return result_validator.validate_existing(
                    outcome, provider=provider
                )

            recognizer = ModeAwareRecognitionExtractor(
                policy=_recognition_policy,
                cloud_extractors={CloudProviderId.GLM: remote},
                local_extractor=local_recognizer,
                artifact_path_resolver=_prepared_artifact_path,
                cloud_result_validator=_validate_cloud_result,
            )

        def _progress(completed, total, percent):
            self.progress = min(90, 45 + int(percent * 0.45))
            self.status_text = f"正在解析发票 ({completed}/{total})..."
            self._safe_emit_stage_event(
                "extraction_pipeline",
                "progress",
                {"completed": completed, "total": total, "percent": percent},
            )

        def _trace(event):
            trace_store.set_fields(event["document_id"], extraction_result=dict(event))
            self._safe_emit_stage_event("extraction_pipeline", "trace", dict(event))

        adapter = AppArchiveAdapter(
            api=self,
            extractor=_extractor,
            save_path=save_path,
            business_records=business_records,
            trace_store=trace_store,
            pairing_finalizer=_pairing_finalizer,
        )
        archive_service = ArchiveService(
            normalizer=adapter.normalize,
            classifier=adapter.classify,
            archive_operation=_archive_operation or adapter.archive_operation,
            dedupe_key=adapter.dedupe_key,
            existing_dedupe_keys=business_records.keys(),
            finalizer=adapter.finalize,
            event_sink=lambda event: self._safe_emit_stage_event(
                "archive_service", "artifact", event
            ),
        )

        pipeline = ExtractionPipeline(
            local_parser=preflight,
            remote_extractor=recognizer,
            max_workers=2,
            verified_ceiling=recognizer.verified_ceiling,
            stop_requested=lambda: bool(getattr(self, "_stop_requested", False)),
            progress_callback=_progress,
            trace_sink=_trace,
        )
        url_recovery_scheduler = DeferredUrlRecoveryScheduler(
            max_workers=DEFAULT_URL_RECOVERY_MAX_WORKERS,
            stop_requested=lambda: bool(getattr(self, "_stop_requested", False)),
            progress_callback=_progress,
        )
        return _ProcessingPipelineSession(
            api=self,
            candidates=candidates,
            pipeline=pipeline,
            archive_service=archive_service,
            save_path=save_path,
            output_state_dir=output_state_dir,
            working_history=working_history,
            business_records=business_records,
            sidecar=sidecar,
            trace_store=trace_store,
            owned_extractor=_owned_extractor,
            provider_retry_delay_seconds=20.0,
            url_recovery_scheduler=url_recovery_scheduler,
        )

    def _retain_artifact(self, save_path, source_path, bucket, reason, metadata=None):
        """在 staging 清理前保留一份可追踪的原件副本。"""
        import json
        import shutil
        import uuid

        retention_dir = os.path.join(save_path, "_audit_retention", bucket)
        os.makedirs(retention_dir, exist_ok=True)
        runtime_metadata = dict(metadata or {})
        safe_metadata = self._sanitize_url_persistence_payload(runtime_metadata)

        is_url_placeholder = (
            runtime_metadata
            and (
                runtime_metadata.get("source_kind") == "url"
                or str(source_path).startswith(("http://", "https://"))
            )
            and (not source_path or not os.path.exists(source_path))
        )
        batched_url_retention = bool(
            is_url_placeholder
            and bucket == "pipeline_retained"
            and str(runtime_metadata.get("candidate_action") or "") == "retain_only"
            and not str(runtime_metadata.get("provider_family") or "").strip()
        )

        if is_url_placeholder:
            url_evidence = build_url_evidence(source_path, bucket)
            if batched_url_retention:
                target_path = os.path.join(retention_dir, "url_retention_index.jsonl")
            else:
                candidate_index = int(runtime_metadata.get("candidate_index", 1) or 1)
                original_name = (
                    f"LinkRetention_{url_evidence['source_hash'][:16]}_"
                    f"{candidate_index}.url.txt"
                )
                target_name = original_name
                target_path = os.path.join(retention_dir, target_name)
                while os.path.exists(target_path):
                    stem, ext = os.path.splitext(original_name)
                    target_name = f"{stem}_{uuid.uuid4().hex[:6]}{ext}"
                    target_path = os.path.join(retention_dir, target_name)
                with open(target_path, "w", encoding="utf-8") as fh:
                    json.dump(url_evidence, fh, ensure_ascii=False, indent=2)
        else:
            if not source_path or not os.path.exists(source_path):
                return source_path

            original_name = os.path.basename(source_path)
            target_name = original_name
            target_path = os.path.join(retention_dir, target_name)
            while os.path.exists(target_path):
                stem, ext = os.path.splitext(original_name)
                target_name = f"{stem}_{uuid.uuid4().hex[:6]}{ext}"
                target_path = os.path.join(retention_dir, target_name)

            shutil.copy2(source_path, target_path)

        payload = {
            "kind": "retention",
            "status": "retained",
            "reason_code": str(bucket),
            "reason_hash": stable_hash(reason),
            "source_hash": stable_hash(
                source_path if is_url_placeholder else self._user_safe_source_reference(source_path)
            ),
            "retained_name_hash": stable_hash(os.path.basename(target_path)),
            "captured_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        }
        if is_url_placeholder:
            payload.update(url_evidence)
        if safe_metadata:
            payload["metadata"] = safe_metadata

        if batched_url_retention:
            with self._url_retention_ledger_lock:
                with open(target_path, "a", encoding="utf-8") as fh:
                    fh.write(json.dumps(payload, ensure_ascii=False, separators=(",", ":")))
                    fh.write("\n")
        else:
            sidecar = f"{target_path}.json"
            try:
                with open(sidecar, "w", encoding="utf-8") as fh:
                    json.dump(payload, fh, ensure_ascii=False, indent=2)
            except Exception as exc:
                print(f"Failed to write retention sidecar {sidecar}: {exc}")

        self.audit_counts["retention"] = int(self.audit_counts.get("retention", 0) or 0) + 1
        self._safe_emit_artifact_event(
            "retention",
            os.path.basename(target_path),
            document_id=safe_metadata.get("document_hash"),
            source_kind=safe_metadata.get("source_kind"),
            reason_code=safe_metadata.get("prefilter_reason_code") or bucket,
            category=bucket,
            extra={
                "bucket": bucket,
                "retention_reason_hash": stable_hash(reason),
                "metadata": safe_metadata,
            },
        )

        return target_path

    def _send_to_manual_check(self, save_path, source_path, reason, metadata=None, is_url=False):
        """把待人工复核候选写入用户输出目录下的中文复核目录。"""
        import json
        import shutil
        import uuid

        manual_dir = os.path.join(save_path, MANUAL_REVIEW_FOLDER)
        os.makedirs(manual_dir, exist_ok=True)
        runtime_metadata = dict(metadata or {})
        safe_metadata = self._sanitize_url_persistence_payload(runtime_metadata)

        def _unique_path(filename):
            target_path = os.path.join(manual_dir, filename)
            while os.path.exists(target_path):
                stem, ext = os.path.splitext(filename)
                filename_local = f"{stem}_{uuid.uuid4().hex[:6]}{ext}"
                target_path = os.path.join(manual_dir, filename_local)
                filename = filename_local
            return target_path

        if is_url:
            url_evidence = build_url_evidence(source_path, reason)
            candidate_index = int(runtime_metadata.get("candidate_index", 1) or 1)
            target_path = _unique_path(
                f"P0_LinkReview_{url_evidence['source_hash'][:16]}_"
                f"{candidate_index}.url.txt"
            )
            with open(target_path, "w", encoding="utf-8") as fh:
                json.dump(url_evidence, fh, ensure_ascii=False, indent=2)
        else:
            if not source_path or not os.path.exists(source_path):
                return source_path
            original_name = os.path.basename(source_path)
            prefix = "P0_Review"
            if runtime_metadata.get("file_name"):
                original_name = os.path.basename(str(runtime_metadata["file_name"]))
            target_path = _unique_path(f"{prefix}_{original_name}")
            shutil.copy2(source_path, target_path)

        sidecar = f"{target_path}.json"
        payload = {
            "kind": "manual_check",
            "status": "pending_review",
            "reason_hash": stable_hash(reason),
            "source_hash": stable_hash(
                source_path if is_url else self._user_safe_source_reference(source_path)
            ),
            "review_name_hash": stable_hash(os.path.basename(target_path)),
            "captured_at": time.strftime("%Y-%m-%d %H:%M:%S"),
            "is_url": is_url,
        }
        if is_url:
            payload.update(url_evidence)
        if safe_metadata:
            payload["metadata"] = safe_metadata

        try:
            with open(sidecar, "w", encoding="utf-8") as fh:
                json.dump(payload, fh, ensure_ascii=False, indent=2)
        except Exception as exc:
            print(f"Failed to write manual-check sidecar {sidecar}: {exc}")

        self.audit_counts["manual_check"] = int(self.audit_counts.get("manual_check", 0) or 0) + 1
        self._safe_emit_artifact_event(
            "manual_check",
            os.path.basename(target_path),
            document_id=safe_metadata.get("document_hash"),
            source_kind=safe_metadata.get("source_kind"),
            reason_code="manual_review",
            category=MANUAL_REVIEW_FOLDER,
            extra={
                "is_url": bool(is_url),
                "reason_hash": stable_hash(reason),
                "metadata": safe_metadata,
            },
        )

        return target_path

    def _cwt_cancellation_matching(self, save_path):
        """CWT 取消撮合: 根据取消知会中的人名，在已归档的住宿确认单中寻找匹配的预订，
        将匹配的预订也移入待人工复核，并在 sidecar 中注明撮合关系。"""
        import re
        import shutil
        import json

        cancellations = getattr(self, '_cwt_cancellation_registry', [])
        if not cancellations:
            return

        # 从取消知会文件名中提取人名
        # 格式: 酒店预定取消知会-{name}-{date}入住-{city} (CONNECT 订单号：{order}).pdf
        cancel_pattern = re.compile(r'取消知会[_\-]?(\S+?)[_\-]\d')

        manual_dir = os.path.join(save_path, MANUAL_REVIEW_FOLDER)
        hotel_dir = os.path.join(save_path, "住宿发票")

        for cancel in cancellations:
            cancel_fn = cancel.get("file_name", "")
            m = cancel_pattern.search(cancel_fn)
            if not m:
                continue
            person_name = m.group(1)
            if not person_name or len(person_name) < 2:
                continue

            # 搜索住宿发票目录中匹配此人名的住宿确认单
            if not os.path.isdir(hotel_dir):
                continue
            for fn in os.listdir(hotel_dir):
                if person_name in fn and "酒店" in fn.lower():
                    src = os.path.join(hotel_dir, fn)
                    dst = os.path.join(manual_dir, f"P0_CancelMatch_{fn}")
                    if os.path.exists(dst):
                        continue
                    os.makedirs(manual_dir, exist_ok=True)
                    shutil.move(src, dst)
                    # 写 sidecar
                    sidecar = {
                        "reason": "CWT_CANCELLATION_MATCH",
                        "matched_cancellation": cancel_fn,
                        "matched_person": person_name,
                        "original_path": self._user_safe_source_reference(src),
                        "review_path": dst,
                        "captured_at": time.strftime("%Y-%m-%d %H:%M:%S"),
                    }
                    try:
                        with open(f"{dst}.json", "w", encoding="utf-8") as fh:
                            json.dump(sidecar, fh, ensure_ascii=False, indent=2)
                    except Exception:
                        pass
                    self.logs.append({"time": time.strftime("[%H:%M:%S]"), "type": "撮合:", "color": "text-amber-400", "msg": f"匹配到取消对应的预订: {fn} ↔ {cancel_fn}"})

    def _record_error_log(
        self,
        save_path,
        email_title,
        error_reason,
        *,
        url_candidate_info=None,
    ):
        """记录错误日志到对应的 csv 文件中"""
        import os
        import csv
        from datetime import datetime
        
        # 目录保护
        if not os.path.exists(save_path):
            os.makedirs(save_path, exist_ok=True)
            
        if url_candidate_info:
            email_title = self._url_candidate_label(url_candidate_info)
            error_reason = f"URL_CANDIDATE_ERROR:{stable_hash(error_reason)[:16]}"

        log_file = os.path.join(save_path, "异常发票处理日志.csv")
        file_exists = os.path.exists(log_file)
        
        try:
            with open(log_file, 'a', newline='', encoding='utf-8-sig') as csvfile:
                fieldnames = ['处理时间', '邮件标题', '错误原因']
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                if not file_exists:
                    writer.writeheader()
                writer.writerow({
                    '处理时间': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    '邮件标题': email_title,
                    '错误原因': error_reason
                })
        except Exception as e:
            print(f"Failed to write error log: {e}")

    def _cleanup_temp_folders(self, staging_dir=None, temp_dir=None):
        """自动清理程序执行过程中产生的临时文件夹"""
        import shutil

        if staging_dir is None and self._active_run_handle is not None:
            staging_dir = self._active_run_handle.staging_dir
        if temp_dir is None:
            temp_dir = self._active_temp_dir
        temp_paths = [staging_dir, temp_dir]
        failures = []

        for target_path in temp_paths:
            if not target_path or not os.path.exists(target_path) or not os.path.isdir(target_path):
                continue
            try:
                shutil.rmtree(target_path)
                print(f"Cleaned up temp folder: {target_path}")
            except Exception as e:
                failure_id = stable_hash(f"{type(e).__name__}:{e}")[:12]
                print(f"Failed to clean temporary folder [{failure_id}]")
                self.logs.append({
                    "time": time.strftime("[%H:%M:%S]"),
                    "type": "ERROR",
                    "color": "text-error",
                    "msg": f"Failed to clean temporary folder [{failure_id}]",
                })
                failures.append((type(e).__name__, failure_id))
        if failures:
            exc_type, failure_id = failures[0]
            raise RuntimeError(f"cleanup failed:{exc_type}:{failure_id}:count={len(failures)}")

    def _legacy_export_result_detail_pre_release_prep(self, export_path=""):
        try:
            from openpyxl import Workbook
        except ImportError:
            return {"success": False, "message": "缺少 openpyxl 依赖，无法导出 Excel"}

        target_dir = export_path or self._effective_save_path or self._requested_save_path or self.get_default_save_path()
        os.makedirs(target_dir, exist_ok=True)

        def _parse_amount(value):
            text = str(value or "").strip()
            if not text:
                return 0.0
            cleaned = re.sub(r"[^\d.\-]", "", text)
            if not cleaned:
                return 0.0
            try:
                return float(cleaned)
            except Exception:
                return 0.0

        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "分类汇总"
        summary_sheet.append(["发票分类", "数量", "金额合计"])

        category_summary = {}
        for item in self.processed_invoices:
            category = str(item.get("category") or "未分类")
            bucket = category_summary.setdefault(category, {"count": 0, "amount": 0.0})
            bucket["count"] += 1
            bucket["amount"] += _parse_amount(item.get("amount", ""))

        if category_summary:
            for category in sorted(category_summary.keys()):
                bucket = category_summary[category]
                summary_sheet.append([category, bucket["count"], round(bucket["amount"], 2)])
        else:
            summary_sheet.append(["暂无成功记录", 0, 0.0])

        success_sheet = workbook.create_sheet("成功明细")
        success_sheet.append(["日期", "金额", "销售方", "分类", "文件路径"])
        for item in self.processed_invoices:
            success_sheet.append([
                item.get("date", ""),
                item.get("amount", ""),
                item.get("merchant", ""),
                item.get("category", ""),
                item.get("path", ""),
            ])

        error_sheet = workbook.create_sheet("异常记录")
        error_sheet.append(["分组", "状态", "原因", "日期", "金额", "销售方", "文件名", "文件路径"])
        for group in self._group_error_invoices():
            for item in group.get("items", []):
                error_sheet.append([
                    group.get("label", ""),
                    item.get("status", ""),
                    item.get("reason", ""),
                    item.get("date", ""),
                    item.get("amount", ""),
                    item.get("merchant", ""),
                    item.get("name", ""),
                    item.get("path", ""),
                ])

        export_file = os.path.join(target_dir, f"结果明细_{time.strftime('%Y%m%d_%H%M%S')}.xlsx")
        workbook.save(export_file)
        self._last_export_path = export_file
        return {"success": True, "message": "结果明细已导出", "path": export_file}


    def _start_async_finalizers(self, fetcher=None):
        handle = self._active_run_handle or self._prepare_run_lifecycle()
        if handle.state not in {RunState.FINALIZING, RunState.COMPLETED, RunState.FAILED}:
            handle.advance(RunState.REPORTING)
        if self.run_state != "finalizing":
            self._mark_finalizing()
        self._safe_emit_stage_event("cleanup_finalize", "enter")
        staging_dir = handle.staging_dir
        temp_dir = self._active_temp_dir
        callbacks = [("report", self._await_truth_audit)]
        if fetcher is not None:
            callbacks.append(("disconnect", fetcher.disconnect))
        callbacks.append(
            (
                "cleanup",
                lambda: self._cleanup_temp_folders(staging_dir=staging_dir, temp_dir=temp_dir),
            )
        )
        state = handle.finalize(callbacks)
        self._safe_emit_stage_event(
            "cleanup_finalize",
            "exit",
            {"result": "failed" if state is RunState.FAILED else "completed"},
        )
        return state

    def _mark_finalizing(self):
        finalizing_progress = self.progress if self.progress >= 95 else 99
        self._set_run_state("finalizing", status_text="正在收尾...", progress=finalizing_progress)

    def _fail_run(
        self,
        status_text,
        error_message,
        fetcher=None,
        include_traceback=False,
        reason_code="",
        user_message="",
    ):
        active_exc = sys.exc_info()[1]
        self._packaged_diag_write(
            "fail_run",
            "_fail_run",
            "exception" if active_exc is not None else "failure",
            summary={"include_traceback": bool(include_traceback)},
            exc=active_exc,
        )
        handle = self._active_run_handle
        if handle is None:
            handle = self._prepare_run_lifecycle()
        if not reason_code or not user_message:
            inferred_code, inferred_message = self._safe_failure_contract(status_text, error_message)
            reason_code = reason_code or inferred_code
            user_message = user_message or inferred_message
        handle.fail(
            active_exc or RuntimeError(error_message or reason_code),
            reason_code=reason_code,
            user_message=user_message,
        )
        self._mark_finalizing()
        self._start_async_finalizers(fetcher)

        self.logs.append({
            "time": time.strftime("[%H:%M:%S]"),
            "type": "ERROR",
            "color": "text-error",
            "msg": f"System exception: {handle.error}",
        })
        self._finish_run(False, status_text, last_error=handle.error)

    def _validate_date_range(self, date_from, date_to):
        date_pattern = re.compile(r"^\d{4}-\d{2}-\d{2}$")
        if date_from and not date_pattern.match(str(date_from)):
            return "开始日期格式必须为 YYYY-MM-DD"
        if date_to and not date_pattern.match(str(date_to)):
            return "结束日期格式必须为 YYYY-MM-DD"
        if date_from and date_to and date_from > date_to:
            return "开始日期不能晚于结束日期"
        return ""

    def _manual_check_path(self):
        base_path = self._effective_save_path or self._requested_save_path or self.get_default_save_path()
        return os.path.join(base_path, MANUAL_REVIEW_FOLDER)

    def _build_error_breakdown(self):
        breakdown = {
            "manual_review": 0,
            "retained_record": 0,
            "processing_error": 0,
        }
        reason_codes = {}

        for item in list(self.error_invoices):
            group_key, _group_label = self._classify_error_invoice(item)
            if group_key not in breakdown:
                group_key = "processing_error"
            breakdown[group_key] += 1
            reason_code = str((item or {}).get("reason", "") or "").strip()
            if reason_code:
                reason_codes[reason_code] = int(reason_codes.get(reason_code, 0) or 0) + 1

        breakdown["reason_codes"] = reason_codes
        return breakdown

    def _summarize_stats(self):
        breakdown = self._build_error_breakdown()
        manual_review_count = max(
            int(breakdown.get("manual_review", 0) or 0),
            int(self.audit_counts.get("manual_check", 0) or 0),
        )
        retention_count = max(
            int(breakdown.get("retained_record", 0) or 0),
            int(self.audit_counts.get("retention", 0) or 0),
        )
        return {
            "emails": int(self.stats.get("emails", 0) or 0),
            "success_count": len(self.processed_invoices),
            "error_count": len(self.error_invoices),
            "manual_check_count": manual_review_count,
            "retention_count": retention_count,
            "raw_invoice_count": int(self.audit_counts.get("raw_invoices", 0) or 0),
            "processing_error_count": int(breakdown.get("processing_error", 0) or 0),
            "result_breakdown": {
                "manual_review": manual_review_count,
                "retained_record": retention_count,
                "processing_error": int(breakdown.get("processing_error", 0) or 0),
            },
            "reason_code_breakdown": breakdown.get("reason_codes", {}),
            "timing_breakdown": self._build_timing_breakdown(),
            "quota_exhausted": bool(self.quota_exhausted),
            "quota_message": self.quota_message,
            "run_state": self.run_state,
            "status_text": self.status_text,
        }

    def _classify_error_invoice(self, item):
        path = str((item or {}).get("path", "") or "")
        artifact_kind = str((item or {}).get("artifact_kind", "") or "").strip().lower()
        status = str((item or {}).get("status", "") or "")
        reason = str((item or {}).get("reason", "") or "")
        category = str((item or {}).get("category", "") or "")
        merchant = str((item or {}).get("merchant", "") or "")
        normalized = " ".join([path, status, reason, category, merchant]).lower()
        normalized_path = path.replace("\\", "/").lower()

        if artifact_kind == "manual_check" or "/manual_check/" in normalized_path or f"/{MANUAL_REVIEW_FOLDER.lower()}/" in normalized_path:
            return "manual_review", "待人工复核"
        if artifact_kind == "retention" or "/_audit_retention/" in normalized_path:
            return "retained_record", "保留记录"
        if "manual_check" in normalized or "人工复核" in normalized or "待人工复核" in normalized:
            return "manual_review", "待人工复核"
        if any(token in normalized for token in [
            "_audit_retention",
            "controlled_run_non_provider_url",
            "history_skipped",
            "prefilter_b_retained",
            "provider_recovery_failed",
            "url_non_invoice_page_skipped",
            "model_rejected",
            "受控跑批保全",
            "已保全待判断",
            "保全",
            "保留",
            "retention",
        ]):
            return "retained_record", "保留记录"
        if any(token in normalized for token in [
            "url_page_timeout",
            "url_auth_wall_detected",
            "url_no_response",
            "url_download_failed",
            "链接下载失败",
        ]):
            return "processing_error", "真实处理异常"
        if self._resolve_quota_message(normalized):
            return "processing_error", "真实处理异常"
        if any(token in normalized for token in [
            "processing_errors",
            "处理中断遗漏",
            "pipeline 断层遗漏",
            "系统遗漏",
            "处理单张票据时抛出异常",
            "异常",
            "failed",
            "error",
        ]):
            return "processing_error", "真实处理异常"
        return "processing_error", "真实处理异常"

    def _group_error_invoices(self):
        grouped = {}
        for item in list(self.error_invoices):
            group_key, group_label = self._classify_error_invoice(item)
            group = grouped.setdefault(group_key, {"key": group_key, "label": group_label, "count": 0, "items": []})
            enriched = dict(item)
            enriched["groupKey"] = group_key
            enriched["groupLabel"] = group_label
            group["items"].append(enriched)
            group["count"] += 1
        return list(grouped.values())

    def stop_processing(self):
        if self.run_state == "finalizing":
            return {"success": False, "message": "任务正在收尾，请稍候"}
        if not self._is_running or not (self._worker_thread and self._worker_thread.is_alive()):
            return {"success": False, "message": "当前没有正在运行的任务"}
        self._request_safe_stop()
        return {"success": True, "message": "已收到停止指令，当前文件处理完成后将结束任务"}

    def close_window(self):
        try:
            import webview
        except ImportError:
            return {"success": False, "message": "桌面窗口接口不可用"}

        if not webview.windows:
            return {"success": False, "message": "当前没有可关闭的桌面窗口"}

        window = webview.windows[0]

        def _shutdown():
            try:
                window.destroy()
            finally:
                time.sleep(0.15)
                os._exit(0)

        threading.Thread(target=_shutdown, daemon=True).start()
        return {"success": True, "message": "窗口关闭中"}

    def _active_desktop_window(self, action_label):
        try:
            import webview
        except ImportError:
            return None, {"success": False, "message": "桌面窗口接口不可用"}

        if not webview.windows:
            return None, {"success": False, "message": f"当前没有可{action_label}的桌面窗口"}
        return webview.windows[0], None

    def minimize_window(self):
        window, error = self._active_desktop_window("最小化")
        if error:
            return error

        try:
            window.minimize()
            return {"success": True, "message": "窗口已最小化"}
        except Exception as exc:
            return {"success": False, "message": f"窗口最小化失败: {exc}"}

    def maximize_window(self):
        window, error = self._active_desktop_window("最大化")
        if error:
            return error

        try:
            window.maximize()
            return {"success": True, "message": "窗口已最大化"}
        except Exception as exc:
            return {"success": False, "message": f"窗口最大化失败: {exc}"}

    def open_manual_check_folder(self):
        manual_path = self._manual_check_path()
        return self.open_folder(manual_path)

    def _legacy_export_run_summary_pre_release_prep(self, export_path=""):
        try:
            from openpyxl import Workbook
        except ImportError:
            return {"success": False, "message": "缺少 openpyxl 依赖，无法导出 Excel"}

        target_dir = export_path or self._effective_save_path or self._requested_save_path or self.get_default_save_path()
        os.makedirs(target_dir, exist_ok=True)

        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "运行汇总"
        summary_sheet.append(["项目", "值"])
        for key, value in self._summarize_stats().items():
            summary_sheet.append([key, value])

        success_sheet = workbook.create_sheet("成功记录")
        success_sheet.append(["日期", "金额", "销售方", "分类", "文件路径"])
        for item in self.processed_invoices:
            success_sheet.append([
                item.get("date", ""),
                item.get("amount", ""),
                item.get("merchant", ""),
                item.get("category", ""),
                item.get("path", ""),
            ])

        error_sheet = workbook.create_sheet("异常记录")
        error_sheet.append(["分组", "状态", "原因", "日期", "金额", "销售方", "文件名", "文件路径"])
        for group in self._group_error_invoices():
            for item in group.get("items", []):
                error_sheet.append([
                    group.get("label", ""),
                    item.get("status", ""),
                    item.get("reason", ""),
                    item.get("date", ""),
                    item.get("amount", ""),
                    item.get("merchant", ""),
                    item.get("name", ""),
                    item.get("path", ""),
                ])

        export_file = os.path.join(target_dir, f"运行摘要_{time.strftime('%Y%m%d_%H%M%S')}.xlsx")
        workbook.save(export_file)
        self._last_export_path = export_file
        return {"success": True, "message": "运行摘要已导出", "path": export_file}

    def _finalize_admission_failure(self, handle, exc, *, reason_code="WORKER_START_FAILED"):
        if handle.state in {RunState.COMPLETED, RunState.FAILED}:
            return handle.state
        handle.fail(
            exc,
            reason_code=reason_code,
            user_message=str(
                getattr(exc, "user_message", "") or "后台任务启动失败，请重试。"
            ),
        )
        self._run_state_store.update(
            run_state="finalizing",
            progress=99,
            status_text="正在收尾...",
        )
        staging_dir = handle.staging_dir
        temp_dir = self._active_temp_dir
        state = handle.finalize(
            [
                ("report", self._await_truth_audit),
                (
                    "cleanup",
                    lambda: self._cleanup_temp_folders(
                        staging_dir=staging_dir,
                        temp_dir=temp_dir,
                    ),
                ),
            ]
        )
        error = handle.error
        self._run_state_store.terminalize(
            "failed",
            status_text="启动失败",
            last_error=error,
            reason_code=reason_code,
            logs=[
                {
                    "time": time.strftime("[%H:%M:%S]"),
                    "type": "ERROR",
                    "color": "text-error",
                    "msg": f"System exception: {error}",
                }
            ],
        )
        self._worker_thread = None
        return state

    def _build_admission_candidate(
        self,
        *,
        rules_text,
        save_path,
        date_from,
        date_to,
        email_address,
        auth_code,
        api_key,
    ):
        run_context = dict(load_run_context() or {})
        requested_save_path = str(
            save_path
            or (
                run_context.get("output_dir", "")
                if run_context.get("enabled")
                else resolve_default_save_path()
            )
        )
        effective_save_path = str(
            run_context.get("output_dir", requested_save_path)
            if run_context.get("enabled")
            else requested_save_path
        )
        effective_date_from = str(
            (run_context.get("locked_date_from", "") if run_context.get("enabled") else "")
            or date_from
            or ""
        )
        effective_date_to = str(
            (run_context.get("locked_date_to", "") if run_context.get("enabled") else "")
            or date_to
            or ""
        )
        date_error = self._validate_date_range(effective_date_from, effective_date_to)
        if date_error:
            return None, None, date_error

        email_text = str(email_address or "")
        email_domain = email_text.rsplit("@", 1)[-1].lower() if "@" in email_text else ""
        candidate = _RunAdmissionCandidate(
            rules_text=str(rules_text or ""),
            requested_save_path=requested_save_path,
            effective_save_path=effective_save_path,
            date_from=effective_date_from,
            date_to=effective_date_to,
            account_id=stable_hash(email_text)[:12],
            channel_id=email_domain.split(".", 1)[0] or "email",
            email_domain=email_domain,
            run_context_json=json.dumps(run_context, ensure_ascii=False, sort_keys=True),
            started_at=time.strftime("%Y-%m-%d %H:%M:%S"),
        )
        return candidate, _RunAdmissionSecrets(email_text, auth_code, api_key), ""

    def _restore_failed_admission(self, previous, previous_settings, settings_existed, settings_touched):
        if settings_touched:
            try:
                if settings_existed:
                    self._settings_store.save(previous_settings)
                else:
                    self._settings_store.clear()
            except Exception:
                pass
        for name, value in previous.items():
            setattr(self, name, value)

    def _reset_run_state_for_admission(self, run_id):
        reset_kwargs = {
            "status_text": "正在准备运行...",
            "run_state": "running",
            "progress": 5,
        }
        try:
            self._run_state_store.reset(run_id, **reset_kwargs)
            return
        except Exception:
            previous_store = self._run_state_store
            self._run_state_store = RunStateStore(
                event_sink=getattr(previous_store, "_event_sink", None),
                state_sink=self._apply_run_state_snapshot,
            )
            try:
                self._run_state_store.reset(run_id, **reset_kwargs)
            except Exception:
                pass
            raise RuntimeError("run state initialization failed") from None

    @staticmethod
    def _admission_missing_directories(context, save_path):
        targets = {
            str(context.get(key) or "").strip()
            for key in (
                "run_root",
                "output_dir",
                "staging_dir",
                "diagnostics_dir",
                "monitoring_dir",
                "qc_dir",
            )
        }
        targets.add(str(save_path or "").strip())
        missing = set()
        for value in targets:
            if not value:
                continue
            path = Path(value).resolve()
            while not path.exists() and path != path.parent:
                missing.add(path)
                path = path.parent
        return tuple(sorted(missing, key=lambda item: len(item.parts), reverse=True))

    @staticmethod
    def _remove_empty_admission_directories(paths):
        import shutil

        for path in paths:
            try:
                if path.is_dir():
                    shutil.rmtree(path)
                elif path.exists():
                    path.unlink()
            except OSError:
                continue

    @staticmethod
    def _close_run_dependencies(dependencies):
        service = getattr(dependencies, "report_service", None)
        close = getattr(service, "close", None)
        if callable(close):
            try:
                close()
            except Exception:
                pass

    def _admit_processing_run(
        self,
        *,
        candidate,
        secrets,
    ):
        with self._admission_lock:
            worker_active = bool(self._worker_thread and self._worker_thread.is_alive())
            active_handle = self._active_run_handle
            handle_active = bool(
                active_handle
                and active_handle.state not in {RunState.COMPLETED, RunState.FAILED}
            )
            if worker_active or handle_active or not self._run_lifecycle.can_begin:
                return {"success": False, "message": "任务已在运行中"}
            if active_handle is not None:
                self._active_run_handle = None

            previous = {
                "_run_context": self._run_context,
                "_current_run_id": self._current_run_id,
                "_requested_save_path": self._requested_save_path,
                "_effective_save_path": self._effective_save_path,
                "_effective_date_from": self._effective_date_from,
                "_effective_date_to": self._effective_date_to,
                "_active_run_config": self._active_run_config,
                "_active_temp_dir": self._active_temp_dir,
                "_truth_audit_thread": self._truth_audit_thread,
                "_truth_audit_job": self._truth_audit_job,
            }
            previous_settings = {}
            settings_existed = os.path.exists(self._settings_store.settings_path)
            settings_touched = False
            handle = None
            dependencies = None
            recognition_policy = None
            admission_missing_dirs = ()
            try:
                self._run_context = candidate.run_context()
                self._current_run_id = str(self._run_context.get("run_id", "") or "")
                admission_missing_dirs = self._admission_missing_directories(
                    self._run_context,
                    candidate.effective_save_path,
                )
                handle = self._prepare_run_lifecycle()
                self._reset_run_state_for_admission(handle.run_id)
                from run_evidence import RevisionUnavailable

                trusted_revision = str(self._revision_resolver() or "").strip().lower()
                if not re.fullmatch(r"[0-9a-f]{40}", trusted_revision):
                    raise RevisionUnavailable()
                ensure_run_context_dirs(self._run_context)
                previous_settings = self._settings_store.load() or {}
                recognition_policy = RecognitionPolicy.from_settings(previous_settings)
                recognition_policy.validate_for_admission(
                    credentials={CloudProviderId.GLM: bool(secrets.api_key)},
                    supported_cloud_providers={CloudProviderId.GLM},
                )
                self._requested_save_path = candidate.requested_save_path
                self._effective_save_path = candidate.effective_save_path
                self._effective_date_from = candidate.date_from
                self._effective_date_to = candidate.date_to

                active_company = str(previous_settings.get("company") or "").strip()
                remember_settings = bool(previous_settings.get("remember_settings", True))
                self._active_run_config = {
                    "company": active_company,
                    "save_path": candidate.effective_save_path,
                    "date_from": candidate.date_from,
                    "date_to": candidate.date_to,
                    "started_at": candidate.started_at,
                    "recognition_mode": recognition_policy.mode.value,
                    "cloud_provider": (
                        recognition_policy.cloud_provider.value
                        if recognition_policy.cloud_provider is not None
                        else ""
                    ),
                }

                os.makedirs(candidate.effective_save_path, exist_ok=True)
                settings_payload = (
                    {
                        "email": secrets.email_address,
                        "auth_code": secrets.auth_code,
                        "api_key": secrets.api_key,
                        "save_path": candidate.effective_save_path,
                        "date_from": candidate.date_from,
                        "date_to": candidate.date_to,
                        "company": active_company,
                        "remember_settings": True,
                        **recognition_policy.to_settings(),
                    }
                    if remember_settings
                    else {"remember_settings": False}
                )
                settings_touched = True
                save_result = self.save_user_settings(settings_payload)
                if not bool((save_result or {}).get("success")):
                    raise RuntimeError("settings persistence failed")

                request = RunRequest(
                    run_id=handle.run_id,
                    date_from=candidate.date_from,
                    date_to=candidate.date_to,
                    save_path=candidate.effective_save_path,
                    rules_text=candidate.rules_text,
                    account_id=candidate.account_id,
                    channel_id=candidate.channel_id,
                    before_exclusive=(
                        datetime.strptime(candidate.date_to, "%Y-%m-%d")
                        + timedelta(days=1)
                    ).strftime("%Y-%m-%d"),
                    account_domain=candidate.email_domain,
                    mailbox="INBOX",
                    target_identifier=active_company,
                    run_mode=str(
                        self._run_context.get("autostart_mode")
                        or ("controlled-run" if self._run_context.get("enabled") else "interactive")
                    ),
                    run_root=str(self._run_context.get("run_root") or ""),
                    evidence_required=bool(
                        self._run_context.get("enabled")
                        and self._run_context.get("run_root")
                    ),
                    candidate_version=str(
                        self._run_context.get("candidate_version") or "source"
                    ),
                    trusted_revision=trusted_revision,
                    validation_required=bool(
                        self._run_context.get("validation_required", False)
                    ),
                    manifest_included_count=int(
                        self._run_context.get("manifest_included_count", 0) or 0
                    ),
                )
                dependencies = self._build_run_dependencies(
                    request,
                    email_address=secrets.email_address,
                    auth_code=secrets.auth_code,
                    api_key=secrets.api_key,
                    recognition_policy=recognition_policy,
                )
                self._safe_write_run_config(
                    secrets.email_address,
                    auth_code=secrets.auth_code,
                    api_key=secrets.api_key,
                    request=request,
                )
            except Exception as exc:
                self._close_run_dependencies(dependencies)
                if handle is not None:
                    reason_code = str(
                        getattr(exc, "reason_code", "") or "WORKER_START_FAILED"
                    )
                    self._finalize_admission_failure(
                        handle,
                        exc,
                        reason_code=reason_code,
                    )
                self._restore_failed_admission(
                    previous,
                    previous_settings,
                    settings_existed,
                    settings_touched,
                )
                self._remove_empty_admission_directories(admission_missing_dirs)
                return {
                    "success": False,
                    "message": (
                        exc.user_message
                        if isinstance(exc, RecognitionPolicyError)
                        else "后台任务启动失败"
                    ),
                }

            self._run_state_store.append_log(
                "信息",
                "前端请求已接收，后台任务正在启动。",
                "text-blue-400",
            )
            self._packaged_diag_reset(
                {
                    "requested_save_path": candidate.requested_save_path,
                    "effective_save_path": candidate.effective_save_path,
                    "date_from": candidate.date_from,
                    "date_to": candidate.date_to,
                    "email_domain": candidate.email_domain,
                    "has_auth_code": bool(secrets.auth_code),
                    "has_api_key": bool(secrets.api_key),
                }
            )
            self._packaged_diag_write(
                "progress_5_written",
                "start_processing",
                "success",
                summary={
                    "requested_save_path": candidate.requested_save_path,
                    "effective_save_path": candidate.effective_save_path,
                    "date_from": candidate.date_from,
                    "date_to": candidate.date_to,
                    "email_domain": candidate.email_domain,
                    "has_auth_code": bool(secrets.auth_code),
                    "has_api_key": bool(secrets.api_key),
                },
            )
            self._safe_emit_stage_event(
                "start_processing",
                "enter",
                {
                    "requested_save_path": candidate.requested_save_path,
                    "effective_save_path": candidate.effective_save_path,
                    "date_from": candidate.date_from,
                    "date_to": candidate.date_to,
                    **self._sensitive_summary(
                        secrets.email_address,
                        secrets.auth_code,
                        secrets.api_key,
                    ),
                },
            )
            try:
                self._start_truth_audit_async(secrets.email_address, secrets.auth_code)
            except Exception as exc:
                self._close_run_dependencies(dependencies)
                self._safe_emit_stage_event(
                    "start_processing",
                    "exit",
                    {"result": "failed", "reason": "WORKER_START_FAILED"},
                )
                self._finalize_admission_failure(handle, exc)
                self._restore_failed_admission(
                    previous,
                    previous_settings,
                    settings_existed,
                    settings_touched,
                )
                return {"success": False, "message": "后台任务启动失败"}

            thread = threading.Thread(
                target=self._processing_worker,
                args=(request, handle, dependencies),
                name="InvoiceFlowWorker",
                daemon=True,
            )
            self._worker_thread = thread
            self._packaged_diag_write("worker_thread_created", "start_processing", "success")
            try:
                thread.start()
            except Exception as exc:
                self._close_run_dependencies(dependencies)
                self._finalize_admission_failure(handle, exc)
                self._restore_failed_admission(
                    previous,
                    previous_settings,
                    settings_existed,
                    settings_touched,
                )
                return {"success": False, "message": "后台任务启动失败"}
            self._packaged_diag_write(
                "worker_thread_started",
                "start_processing",
                "success",
                summary={"thread_is_alive": bool(thread.is_alive())},
            )
            self._safe_emit_stage_event("start_processing", "exit", {"result": "started"})
            return {"success": True, "message": "任务已启动"}

    def start_processing(self, rules_text, save_path, date_from=None, date_to=None, email_address=None, auth_code=None, api_key=None):
        if not self._run_lifecycle.can_begin or self._is_running or (self._worker_thread and self._worker_thread.is_alive()):
            return {"success": False, "message": "任务已在运行中"}

        if not email_address or not auth_code:
            return {"success": False, "message": "缺少必要凭证，请填写邮箱和授权码"}

        candidate, secrets, date_error = self._build_admission_candidate(
            rules_text=rules_text,
            save_path=save_path,
            date_from=date_from,
            date_to=date_to,
            email_address=email_address,
            auth_code=auth_code,
            api_key=api_key,
        )
        if date_error:
            return {"success": False, "message": date_error}

        return self._admit_processing_run(
            candidate=candidate,
            secrets=secrets,
        )

    def get_processed_records(self):
        """前端数据分析页面调用，获取已处理的账单或发票记录"""
        self._sync_run_state_store_from_legacy()
        return self._run_state_store.snapshot()["processed_invoices"]

    def get_progress(self):
        """前端轮询进度条和日志调用"""
        self._sync_run_state_store_from_legacy()
        payload = self._run_state_store.frontend_snapshot(
            build_identity=self.build_identity,
            raw_date_range=self._raw_date_range_display,
            imap_query_range=self._imap_query_range_display,
        )
        self._packaged_diag_log_progress_poll(payload)
        return payload

    def get_results(self):
        """前端分析页调用，获取最终的统计数据"""
        self._sync_run_state_store_from_legacy()
        state_snapshot = self._run_state_store.snapshot()
        processed_invoices = state_snapshot["processed_invoices"]
        error_invoices = state_snapshot["error_invoices"]
        categories = state_snapshot["new_categories"]
        grouped_errors = self._group_error_invoices()
        summary = self._summarize_stats()
        return {
            "categories": categories,
            "successInvoices": processed_invoices,
            "errorInvoices": error_invoices,
            "groupedErrorInvoices": grouped_errors,
            "manual_check_path": self._manual_check_path(),
            "output_path": self._effective_save_path or self._requested_save_path or self.get_default_save_path(),
            "summary": summary,
            "build_identity": self.build_identity,
            "raw_date_range": self._raw_date_range_display,
            "imap_query_range": self._imap_query_range_display,
            "resultBreakdown": summary.get("result_breakdown", {}),
            "reasonCodeBreakdown": summary.get("reason_code_breakdown", {}),
            "quota_exhausted": self.quota_exhausted,
            "quota_message": self.quota_message,
            "last_export_path": self._last_export_path,
            "invoices": processed_invoices # 兼容旧的数据结构
        }

    def choose_directory(self):
        """调用系统原生目录选择器"""
        run_context = self._refresh_run_context()
        if run_context.get("enabled"):
            return {"success": True, "path": run_context.get("output_dir", "")}
        print("Opening directory dialog...")
        import webview
        # webview.windows 列表包含了所有当前活动的窗口，取第一个
        if webview.windows:
            window = webview.windows[0]
            # webview.FOLDER_DIALOG
            result = window.create_file_dialog(webview.FOLDER_DIALOG)
            if result and len(result) > 0:
                print(f"Selected: {result[0]}")
                return {"success": True, "path": result[0]}
        return {"success": False}

    def open_folder(self, path):
        """调用系统资源管理器打开特定文件夹"""
        import os
        import platform
        import subprocess

        # 确保存储的路径分隔符是环境兼容的
        path = os.path.normpath(path)
        print(f"Opening folder: {path}")

        if not os.path.exists(path):
            try:
                os.makedirs(path, exist_ok=True)
            except Exception as e:
                 return {"success": False, "message": f"无法创建目录: {str(e)}"}

        try:
            if platform.system() == "Windows":
                # Windows：推荐采用 subprocess.Popen 或者 os.startfile 这里采用 explorer 确保前台弹出
                subprocess.Popen(f'explorer "{path}"')
            elif platform.system() == "Darwin":
                # macOS
                subprocess.Popen(["open", path])
            else:
                # Linux
                subprocess.Popen(["xdg-open", path])
            return {"success": True}
        except Exception as e:
            return {"success": False, "message": str(e)}

    def retry_all_errors(self):
        """重新处理所有失败的发票任务（占位实现）"""
        print("Retrying all failed invoices...")
        return {"success": True, "message": "已为您触发全部重试任务"}

    def retry_single_invoice(self, name):
        """重新处理特定的一张发票（占位实现）"""
        print(f"Retrying single invoice: {name}")
        return {"success": True, "message": f"正在重试发票: {name}"}

    def view_invoice(self, path):
        """在系统默认查看器中打开单张发票或图片"""
        print(f"Viewing invoice: {path}")
        import os
        import platform
        import subprocess
        
        path = os.path.normpath(path)
        if not os.path.exists(path):
            return {"success": False, "message": "文件不存在"}
            
        try:
            if platform.system() == "Windows":
                os.startfile(path)
            elif platform.system() == "Darwin":
                subprocess.Popen(["open", path])
            else:
                subprocess.Popen(["xdg-open", path])
            return {"success": True, "message": "已在外部程序中打开文件"}
        except Exception as e:
            return {"success": False, "message": f"打开失败: {str(e)}"}

    def export_run_summary(self, export_path=""):
        try:
            from openpyxl import Workbook
        except ImportError:
            return {"success": False, "message": "缺少 openpyxl 依赖，无法导出 Excel"}

        target_dir = export_path or self._effective_save_path or self._requested_save_path or self.get_default_save_path()
        os.makedirs(target_dir, exist_ok=True)

        def _parse_amount(value):
            text = str(value or "").strip()
            if not text:
                return 0.0
            cleaned = re.sub(r"[^\d.\-]", "", text)
            if not cleaned:
                return 0.0
            try:
                return float(cleaned)
            except Exception:
                return 0.0

        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "分类汇总"
        summary_sheet.append(["发票分类", "数量", "金额合计"])

        category_summary = {}
        for item in self.processed_invoices:
            category = str(item.get("category") or "未分类")
            bucket = category_summary.setdefault(category, {"count": 0, "amount": 0.0})
            bucket["count"] += 1
            bucket["amount"] += _parse_amount(item.get("amount", ""))

        if category_summary:
            for category in sorted(category_summary.keys()):
                bucket = category_summary[category]
                summary_sheet.append([category, bucket["count"], round(bucket["amount"], 2)])
        else:
            summary_sheet.append(["暂无成功记录", 0, 0.0])

        success_sheet = workbook.create_sheet("成功明细")
        success_sheet.append(["日期", "金额", "销售方", "分类", "文件路径"])
        for item in self.processed_invoices:
            success_sheet.append([
                item.get("date", ""),
                item.get("amount", ""),
                item.get("merchant", ""),
                item.get("category", ""),
                item.get("path", ""),
            ])

        error_sheet = workbook.create_sheet("异常记录")
        error_sheet.append(["分组", "状态", "原因", "日期", "金额", "销售方", "文件名", "文件路径"])
        for group in self._group_error_invoices():
            for item in group.get("items", []):
                error_sheet.append([
                    group.get("label", ""),
                    item.get("status", ""),
                    item.get("reason", ""),
                    item.get("date", ""),
                    item.get("amount", ""),
                    item.get("merchant", ""),
                    item.get("name", ""),
                    item.get("path", ""),
                ])

        export_file = os.path.join(target_dir, f"结果明细_{time.strftime('%Y%m%d_%H%M%S')}.xlsx")
        workbook.save(export_file)
        self._last_export_path = export_file
        return {"success": True, "message": "结果明细已导出", "path": export_file}

if __name__ == "__main__":
    # 提供 CLI 测试入口
    print(">>> [测试入口] 从终端直接启动 app_api.py")
    app = InvoiceAppAPI()
    try:
        import time
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        print(">>> 退出程序。")
